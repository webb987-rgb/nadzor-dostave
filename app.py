import re
import time
import json
import random
import datetime
import smtplib
import threading
import requests
import pandas as pd
import streamlit as st
import sqlite3
import os
from concurrent.futures import ThreadPoolExecutor, as_completed
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path
from dotenv import load_dotenv

load_dotenv()  # učitava varijable iz .env fajla

# ─────────────────────────── KONFIGURACIJA ───────────────────────────────────

EMAIL_SENDER   = os.getenv("EMAIL_SENDER", "ampromoscript@gmail.com")
EMAIL_PASSWORD = os.getenv("EMAIL_PASSWORD", "")
SETTINGS_PASSWORD  = os.getenv("SETTINGS_PASSWORD", "zekapeka")
RESET_PASSWORD_ENV = os.getenv("RESET_PASSWORD", "zekapeka")

DB_FILE = Path("promo_monitor.db")

CITY_KEYS    = [
    "Beograd", "Novi Sad", "Nis", "Kragujevac",
    "Arandelovac", "Bor", "Borca", "Cacak", "Jagodina",
    "Kraljevo", "Krusevac", "Lazarevac", "Leskovac",
    "Novi Pazar", "Obrenovac", "Pancevo", "Pozarevac",
    "Smederevo", "Sombor", "Subotica", "Uzice",
    "Valjevo", "Vrsac", "Zlatibor", "Zrenjanin",
]
CITY_DISPLAY = {
    "Beograd":     "Beograd",
    "Novi Sad":    "Novi Sad",
    "Nis":         "Niš",
    "Kragujevac":  "Kragujevac",
    "Arandelovac": "Aranđelovac",
    "Bor":         "Bor",
    "Borca":       "Borča",
    "Cacak":       "Čačak",
    "Jagodina":    "Jagodina",
    "Kraljevo":    "Kraljevo",
    "Krusevac":    "Kruševac",
    "Lazarevac":   "Lazarevac",
    "Leskovac":    "Leskovac",
    "Novi Pazar":  "Novi Pazar",
    "Obrenovac":   "Obrenovac",
    "Pancevo":     "Pančevo",
    "Pozarevac":   "Požarevac",
    "Smederevo":   "Smederevo",
    "Sombor":      "Sombor",
    "Subotica":    "Subotica",
    "Uzice":       "Užice",
    "Valjevo":     "Valjevo",
    "Vrsac":       "Vršac",
    "Zlatibor":    "Zlatibor",
    "Zrenjanin":   "Zrenjanin",
}
CITIES = [CITY_DISPLAY[k] for k in CITY_KEYS]

FETCH_WORKERS = 2          # max simultanih promo HTTP zahteva PO gradu
CITY_PARALLEL = 3          # max gradova koji se skeniraju istovremeno

# Globalni semafor — ograničava UKUPAN broj simultanih HTTP zahteva prema Woltu,
# bez obzira koliko gradova ide paralelno. Tako Wolt vidi isti saobraćaj kao pre.
_global_http_sem = threading.Semaphore(FETCH_WORKERS * CITY_PARALLEL)

EMAIL_IGNORE_PROMOS = [
    "0 din delivery fee for 14 days",
    "0 din delivery fee",
    "free delivery for 14 days",
    "besplatna dostava 14 dana",
    "besplatna dostava",
]

AMM_COLS   = ["restaurant_norm", "restaurant_display", "city", "am_name", "am_email"]
ALERT_COLS = ["timestamp", "city", "restaurant_display", "am_name", "am_email", "akcije"]

SCAN_FILE           = Path("scan_baza_item.json")  # legacy — čuva se samo radi kompatibilnosti pri resetu
LOCK_FILE           = Path("_scan_running.lock")   # legacy — lock je sada u SQLite
COOLDOWN_DAYS       = 7

# ── Višestruke lokacije po gradu ─────────────────────────────────────────────
CITY_MULTI_COORDS = {
    # Beograd – 25 tačaka (originalne + više pokrivanja)
    "Beograd": [
        (44.8610, 20.3450), (44.8395, 20.3662), (44.8251, 20.4102), (44.8130, 20.4182), (44.8050, 20.3880),
        (44.8255, 20.4571), (44.8180, 20.4522), (44.8160, 20.4735), (44.8042, 20.4521), (44.8180, 20.4620),
        (44.8001, 20.4705), (44.8145, 20.4990), (44.8080, 20.4905), (44.7932, 20.4800), (44.8175, 20.5182),
        (44.8160, 20.4950), (44.8100, 20.5100), (44.7925, 20.4430), (44.7920, 20.4350), (44.7820, 20.4550),
        (44.7760, 20.4180), (44.7500, 20.4100), (44.7870, 20.4660), (44.7975, 20.4650), (44.8070, 20.4100),
    ],
    # Novi Sad – prošireno na 10
    "Novi Sad": [
        (45.2671, 19.8335), (45.2500, 19.8100), (45.2850, 19.8600), (45.2400, 19.8700), (45.2900, 19.7900),
        (45.2750, 19.8450), (45.2600, 19.8650), (45.2450, 19.8200), (45.2550, 19.7950), (45.2350, 19.8500),
    ],
    # Niš – prošireno na 10
    "Nis": [
        (43.3209, 21.8958), (43.3050, 21.8800), (43.3350, 21.9150), (43.3100, 21.9300), (43.2950, 21.8700),
        (43.3280, 21.9050), (43.3150, 21.8650), (43.3400, 21.8900), (43.3000, 21.9100), (43.3450, 21.9250),
    ],
    # Kragujevac – prošireno na 10
    "Kragujevac": [
        (44.0128, 20.9114), (44.0000, 20.8900), (44.0300, 20.9300), (43.9900, 20.9400),
        (44.0200, 20.9000), (44.0050, 20.9250), (44.0350, 20.9100), (43.9950, 20.9000),
        (44.0150, 20.9500), (44.0280, 20.8800),
    ],
    # Aranđelovac
    "Arandelovac": [
        (44.3028, 20.5611), (44.2950, 20.5500), (44.3100, 20.5700), (44.2880, 20.5750),
        (44.3150, 20.5450), (44.3050, 20.5800), (44.2920, 20.5350), (44.3200, 20.5600),
        (44.2980, 20.5250), (44.3080, 20.5950),
    ],
    # Bor
    "Bor": [
        (44.0769, 22.0958), (44.0650, 22.0800), (44.0900, 22.1100), (44.0580, 22.1050),
        (44.0850, 22.0700), (44.0700, 22.1200), (44.0950, 22.0900), (44.0620, 22.0650),
        (44.0780, 22.1300), (44.1000, 22.0800),
    ],
    # Borča (deo Beograda, Palilula leva obala)
    "Borca": [
        (44.8820, 20.5350), (44.8750, 20.5200), (44.8900, 20.5500), (44.8680, 20.5450),
        (44.8950, 20.5150), (44.8700, 20.5600), (44.8830, 20.5050), (44.8780, 20.5700),
        (44.8640, 20.5300), (44.8920, 20.5400),
    ],
    # Čačak
    "Cacak": [
        (43.8914, 20.3496), (43.8800, 20.3350), (43.9050, 20.3650), (43.8700, 20.3600),
        (43.9100, 20.3300), (43.8850, 20.3750), (43.8750, 20.3200), (43.9000, 20.3500),
        (43.8650, 20.3450), (43.9150, 20.3700),
    ],
    # Jagodina
    "Jagodina": [
        (43.9766, 21.2614), (43.9650, 21.2480), (43.9900, 21.2750), (43.9580, 21.2700),
        (43.9850, 21.2450), (43.9700, 21.2800), (43.9950, 21.2550), (43.9620, 21.2550),
        (43.9780, 21.2900), (43.9530, 21.2650),
    ],
    # Kraljevo
    "Kraljevo": [
        (43.7236, 20.6894), (43.7120, 20.6750), (43.7350, 20.7050), (43.7050, 20.7000),
        (43.7400, 20.6700), (43.7180, 20.7150), (43.7280, 20.6600), (43.7450, 20.6950),
        (43.7000, 20.6850), (43.7320, 20.7200),
    ],
    # Kruševac
    "Krusevac": [
        (43.5833, 21.3333), (43.5700, 21.3200), (43.5980, 21.3480), (43.5620, 21.3450),
        (43.6050, 21.3150), (43.5750, 21.3550), (43.5850, 21.3050), (43.6000, 21.3400),
        (43.5550, 21.3350), (43.5900, 21.3650),
    ],
    # Lazarevac
    "Lazarevac": [
        (44.3800, 20.2569), (44.3680, 20.2430), (44.3930, 20.2700), (44.3620, 20.2650),
        (44.3980, 20.2350), (44.3720, 20.2800), (44.3850, 20.2250), (44.4000, 20.2600),
        (44.3580, 20.2550), (44.3920, 20.2900),
    ],
    # Leskovac
    "Leskovac": [
        (42.9981, 21.9461), (42.9850, 21.9320), (43.0120, 21.9600), (42.9780, 21.9550),
        (43.0180, 21.9300), (42.9920, 21.9680), (43.0050, 21.9200), (43.0200, 21.9500),
        (42.9700, 21.9450), (43.0100, 21.9750),
    ],
    # Novi Pazar
    "Novi Pazar": [
        (43.1367, 20.5122), (43.1250, 20.4980), (43.1500, 20.5280), (43.1180, 20.5200),
        (43.1550, 20.5000), (43.1300, 20.5400), (43.1420, 20.4880), (43.1600, 20.5250),
        (43.1120, 20.5100), (43.1450, 20.5500),
    ],
    # Obrenovac
    "Obrenovac": [
        (44.6547, 20.2111), (44.6430, 20.1980), (44.6680, 20.2250), (44.6360, 20.2200),
        (44.6750, 20.2000), (44.6480, 20.2350), (44.6580, 20.1880), (44.6820, 20.2150),
        (44.6300, 20.2100), (44.6650, 20.2450),
    ],
    # Pančevo
    "Pancevo": [
        (44.8708, 20.6408), (44.8580, 20.6260), (44.8850, 20.6560), (44.8500, 20.6500),
        (44.8920, 20.6300), (44.8640, 20.6650), (44.8750, 20.6150), (44.8980, 20.6480),
        (44.8450, 20.6400), (44.8800, 20.6750),
    ],
    # Požarevac
    "Pozarevac": [
        (44.6197, 21.1869), (44.6080, 21.1720), (44.6330, 21.2020), (44.6010, 21.2000),
        (44.6400, 21.1750), (44.6150, 21.2150), (44.6250, 21.1600), (44.6450, 21.1950),
        (44.5950, 21.1900), (44.6300, 21.2250),
    ],
    # Smederevo
    "Smederevo": [
        (44.6644, 20.9278), (44.6520, 20.9130), (44.6780, 20.9430), (44.6450, 20.9380),
        (44.6850, 20.9180), (44.6580, 20.9550), (44.6700, 20.9050), (44.6920, 20.9350),
        (44.6380, 20.9280), (44.6750, 20.9650),
    ],
    # Sombor
    "Sombor": [
        (45.7772, 19.1122), (45.7650, 19.0980), (45.7900, 19.1280), (45.7580, 19.1200),
        (45.7980, 19.1000), (45.7720, 19.1400), (45.7850, 19.0880), (45.8050, 19.1200),
        (45.7520, 19.1100), (45.7920, 19.1500),
    ],
    # Subotica
    "Subotica": [
        (46.1003, 19.6675), (46.0880, 19.6520), (46.1150, 19.6840), (46.0800, 19.6800),
        (46.1250, 19.6600), (46.0950, 19.6950), (46.1050, 19.6400), (46.1300, 19.6750),
        (46.0750, 19.6700), (46.1150, 19.7050),
    ],
    # Užice
    "Uzice": [
        (43.8567, 19.8483), (43.8450, 19.8340), (43.8700, 19.8640), (43.8380, 19.8600),
        (43.8780, 19.8350), (43.8500, 19.8750), (43.8620, 19.8220), (43.8850, 19.8550),
        (43.8320, 19.8500), (43.8700, 19.8850),
    ],
    # Valjevo
    "Valjevo": [
        (44.2742, 19.8878), (44.2620, 19.8730), (44.2880, 19.9040), (44.2550, 19.9000),
        (44.2950, 19.8780), (44.2700, 19.9150), (44.2800, 19.8630), (44.3000, 19.8980),
        (44.2500, 19.8900), (44.2850, 19.9250),
    ],
    # Vršac
    "Vrsac": [
        (45.1167, 21.3000), (45.1050, 21.2860), (45.1300, 21.3150), (45.0980, 21.3100),
        (45.1380, 21.2900), (45.1120, 21.3280), (45.1230, 21.2750), (45.1450, 21.3050),
        (45.0920, 21.3000), (45.1300, 21.3380),
    ],
    # Zlatibor
    "Zlatibor": [
        (43.7253, 19.7036), (43.7150, 19.6900), (43.7380, 19.7180), (43.7080, 19.7150),
        (43.7450, 19.6980), (43.7200, 19.7300), (43.7300, 19.6800), (43.7500, 19.7100),
        (43.7050, 19.7050), (43.7350, 19.7400),
    ],
    # Zrenjanin
    "Zrenjanin": [
        (45.3819, 20.3833), (45.3700, 20.3690), (45.3950, 20.3980), (45.3630, 20.3930),
        (45.4020, 20.3750), (45.3780, 20.4100), (45.3850, 20.3600), (45.4100, 20.3900),
        (45.3580, 20.3830), (45.3950, 20.4200),
    ],
}

CITY_COORDS = {k: v[0] for k, v in CITY_MULTI_COORDS.items()}

def get_active_coords() -> dict:
    if "custom_coords" in st.session_state:
        return st.session_state["custom_coords"]
    return CITY_MULTI_COORDS

# ─────────────────────────── PAGE CONFIG ─────────────────────────────────────

st.set_page_config(page_title="Promo Monitor", page_icon="🏷️", layout="wide")

st.markdown("""
<style>
    [data-testid="stAppViewContainer"] { background: #f7f8fc; }
    .kpi { background:#fff; border-radius:12px; padding:18px 24px;
           box-shadow:0 2px 8px rgba(0,0,0,0.07); text-align:center; }
    .kpi-val { font-size:2.2rem; font-weight:800; color:#009de0; }
    .kpi-lbl { font-size:.85rem; color:#888; margin-top:4px; }
    div[data-testid="stDataFrame"] thead th { background:#009de0!important; color:#fff!important; }
    .timer-box { font-size:1.1rem; font-weight:700; color:#009de0; padding:6px 16px;
                 background:#e8f6fd; border-radius:8px; display:inline-block; margin-bottom:8px; }
</style>
""", unsafe_allow_html=True)

# ─────────────────────────── LOCAL SQLITE INITIALIZATION ─────────────────────

def init_db():
    with sqlite3.connect(DB_FILE, check_same_thread=False) as conn:
        conn.execute("CREATE TABLE IF NOT EXISTS scan_baza (grad TEXT, naziv TEXT, slug TEXT, status TEXT, ocena TEXT, dostava TEXT, novo TEXT, akcije TEXT, link TEXT, naziv_norm TEXT)")
        conn.execute("CREATE TABLE IF NOT EXISTS amm_baza (restaurant_norm TEXT, restaurant_display TEXT, city TEXT, am_name TEXT, am_email TEXT)")
        conn.execute("CREATE TABLE IF NOT EXISTS alert_log (timestamp TEXT, city TEXT, restaurant_display TEXT, am_name TEXT, am_email TEXT, akcije TEXT)")
        conn.execute("CREATE TABLE IF NOT EXISTS sales_baza (city TEXT, emails TEXT)")
        conn.execute("""CREATE TABLE IF NOT EXISTS promo_state (
            restaurant_norm TEXT,
            city            TEXT,
            last_akcije     TEXT,
            PRIMARY KEY (restaurant_norm, city)
        )""")
        # Istorija promo promena — nova tabela
        conn.execute("""CREATE TABLE IF NOT EXISTS promo_history (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            timestamp       TEXT,
            restaurant_norm TEXT,
            restaurant_display TEXT,
            city            TEXT,
            stare_akcije    TEXT,
            nove_akcije     TEXT
        )""")
        # Alert cooldown — prebačeno iz JSON fajla u SQLite
        conn.execute("""CREATE TABLE IF NOT EXISTS alert_cooldown (
            am_email        TEXT,
            restaurant_norm TEXT,
            last_sent_date  TEXT,
            PRIMARY KEY (am_email, restaurant_norm)
        )""")
        # Sent new restaurants — prebačeno iz JSON fajla u SQLite
        conn.execute("""CREATE TABLE IF NOT EXISTS sent_new_restaurants (
            slug TEXT PRIMARY KEY
        )""")
        conn.execute("""CREATE TABLE IF NOT EXISTS watchlist (
            slug        TEXT,
            city        TEXT,
            naziv       TEXT,
            added_at    TEXT,
            PRIMARY KEY (slug, city)
        )""")

init_db()

# ─────────────────────────── TRACKING (UKLONJENO) ───────────────────────────
# Tracking pixel i mail open-tracking su potpuno uklonjeni.

def tracking_pixel_html(mail_id: str) -> str:
    return ""  # stub — ne radi ništa, kompatibilnost sa email funkcijama

def start_tracking_server_thread():
    pass  # uklonjeno

# ─────────────────────────── HELPERS ─────────────────────────────────────────

def normalize(name: str) -> str:
    return re.sub(r"[^\w]", "", str(name).lower())

def local_now() -> str:
    return datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

def display_to_key(display_name: str) -> str:
    for key, disp in CITY_DISPLAY.items():
        if disp == display_name or key == display_name:
            return key
    norm = normalize(display_name)
    for key in CITY_KEYS:
        if normalize(key) == norm:
            return key
    return display_name

def is_ignored_promo(text: str) -> bool:
    t = text.lower().strip().lstrip("•").strip()
    t = t.replace("[wolt+]", "").strip()
    for ignored in EMAIL_IGNORE_PROMOS:
        if ignored.lower() in t:
            return True
    return False

def filter_akcije_for_email(akcije_str: str) -> str:
    if not akcije_str or akcije_str.strip() == "-":
        return "-"
    lines = [l for l in akcije_str.split("\n") if l.strip()]
    filtered = [l for l in lines if not is_ignored_promo(l)]
    return "\n".join(filtered) if filtered else "-"

# ─────────────────────────── SQLITE PERSISTENTNA BAZA ────────────────────────

_CACHE_KEYS = {
    "amm":   "_cache_amm",
    "alert": "_cache_alert",
    "sales": "_cache_sales",
    "scan":  "_cache_scan",
}

def _cache_get(key: str):
    return st.session_state.get(_CACHE_KEYS[key])

def _cache_set(key: str, value):
    st.session_state[_CACHE_KEYS[key]] = value

# ── SCAN ──────────────────────────────────────────────────────────────────────

def save_scan_gsheet(df: pd.DataFrame):
    try:
        with sqlite3.connect(DB_FILE, check_same_thread=False) as conn:
            df.to_sql("scan_baza", conn, if_exists="replace", index=False)
    except Exception as e:
        st.warning(f"Baza scan save greška: {e}")
    _cache_set("scan", df)

def load_scan_gsheet() -> pd.DataFrame:
    cached = _cache_get("scan")
    if cached is not None:
        return cached
    try:
        with sqlite3.connect(DB_FILE, check_same_thread=False) as conn:
            df = pd.read_sql_query("SELECT * FROM scan_baza", conn)
            if not df.empty:
                _cache_set("scan", df)
                return df
    except Exception:
        pass
    return pd.DataFrame()

def scan_meta_gsheet() -> str | None:
    cached = _cache_get("scan")
    if cached is not None and not cached.empty:
        return "dostupan (keš)"
    try:
        with sqlite3.connect(DB_FILE, check_same_thread=False) as conn:
            row = conn.execute("SELECT MAX(rowid) FROM scan_baza").fetchone()
            if row and row[0]:
                return "SQLite baza"
    except Exception:
        pass
    return None

# ── AMM ───────────────────────────────────────────────────────────────────────

def save_amm_gsheet(df: pd.DataFrame):
    try:
        with sqlite3.connect(DB_FILE, check_same_thread=False) as conn:
            df.to_sql("amm_baza", conn, if_exists="replace", index=False)
    except Exception as e:
        st.warning(f"AMM save greška: {e}")
    _cache_set("amm", df)

def load_amm_gsheet() -> pd.DataFrame:
    cached = _cache_get("amm")
    if cached is not None:
        return cached
    try:
        with sqlite3.connect(DB_FILE, check_same_thread=False) as conn:
            df = pd.read_sql_query("SELECT * FROM amm_baza", conn)
            if not df.empty:
                for c in AMM_COLS:
                    if c not in df.columns:
                        df[c] = ""
                _cache_set("amm", df)
                return df
    except Exception:
        pass
    empty = pd.DataFrame(columns=AMM_COLS)
    _cache_set("amm", empty)
    return empty

# ── ALERT LOG ─────────────────────────────────────────────────────────────────

def save_alert_log_gsheet(df: pd.DataFrame):
    try:
        with sqlite3.connect(DB_FILE, check_same_thread=False) as conn:
            df.to_sql("alert_log", conn, if_exists="replace", index=False)
    except Exception as e:
        st.warning(f"Alert log save greška: {e}")
    _cache_set("alert", df)

def load_alert_log_gsheet() -> pd.DataFrame:
    cached = _cache_get("alert")
    if cached is not None:
        return cached
    try:
        with sqlite3.connect(DB_FILE, check_same_thread=False) as conn:
            df = pd.read_sql_query("SELECT * FROM alert_log", conn)
            if not df.empty:
                for c in ALERT_COLS:
                    if c not in df.columns:
                        df[c] = ""
                _cache_set("alert", df)
                return df
    except Exception:
        pass
    empty = pd.DataFrame(columns=ALERT_COLS)
    _cache_set("alert", empty)
    return empty

def append_alert_log_gsheet(rows: list):
    try:
        new_df = pd.DataFrame(rows)
        with sqlite3.connect(DB_FILE, check_same_thread=False) as conn:
            new_df.to_sql("alert_log", conn, if_exists="append", index=False)
    except Exception as e:
        st.warning(f"Alert append greška: {e}")
    existing_df = _cache_get("alert")
    if existing_df is None:
        existing_df = pd.DataFrame(columns=ALERT_COLS)
    merged = pd.concat([existing_df, pd.DataFrame(rows)], ignore_index=True)
    _cache_set("alert", merged)

# ── SALES ─────────────────────────────────────────────────────────────────────

def load_sales_gsheet() -> dict:
    cached = _cache_get("sales")
    if cached is not None:
        return cached
    try:
        with sqlite3.connect(DB_FILE, check_same_thread=False) as conn:
            df = pd.read_sql_query("SELECT * FROM sales_baza", conn)
            if not df.empty:
                result = {}
                for _, row in df.iterrows():
                    city = row.get("city", "")
                    emails_str = row.get("emails", "")
                    if city:
                        result[city] = [e.strip() for e in emails_str.split(",") if e.strip()]
                _cache_set("sales", result)
                return result
    except Exception:
        pass
    default = {city: [] for city in CITIES}
    _cache_set("sales", default)
    return default

def save_sales_gsheet(data: dict):
    try:
        rows = [{"city": city, "emails": ", ".join(emails)} for city, emails in data.items()]
        df = pd.DataFrame(rows)
        with sqlite3.connect(DB_FILE, check_same_thread=False) as conn:
            df.to_sql("sales_baza", conn, if_exists="replace", index=False)
    except Exception as e:
        st.warning(f"Sales save greška: {e}")
    _cache_set("sales", data)

# ── PROMO STATE (pamćenje akcija po restoranu za pametno slanje AM maila) ─────

def load_promo_state() -> dict:
    """Vraća dict: { (restaurant_norm, city): last_akcije_str }"""
    try:
        with sqlite3.connect(DB_FILE, check_same_thread=False) as conn:
            rows = conn.execute("SELECT restaurant_norm, city, last_akcije FROM promo_state").fetchall()
            return {(r[0], r[1]): r[2] for r in rows}
    except Exception:
        return {}

def save_promo_state_bulk(updates: dict):
    """updates: { (restaurant_norm, city): akcije_str }"""
    try:
        with sqlite3.connect(DB_FILE, check_same_thread=False) as conn:
            conn.executemany(
                "INSERT INTO promo_state (restaurant_norm, city, last_akcije) VALUES (?,?,?) "
                "ON CONFLICT(restaurant_norm, city) DO UPDATE SET last_akcije=excluded.last_akcije",
                [(k[0], k[1], v) for k, v in updates.items()]
            )
    except Exception as e:
        import logging as _log
        _log.getLogger("scheduler").error(f"save_promo_state_bulk greška: {e}")

def save_promo_history(changes: list):
    """
    Snima promene akcija u promo_history tabelu.
    changes: list of dicts: {restaurant_norm, restaurant_display, city, stare_akcije, nove_akcije}
    """
    if not changes:
        return
    try:
        now_str = local_now()
        with sqlite3.connect(DB_FILE, check_same_thread=False) as conn:
            conn.executemany(
                "INSERT INTO promo_history (timestamp, restaurant_norm, restaurant_display, city, stare_akcije, nove_akcije) "
                "VALUES (?,?,?,?,?,?)",
                [(now_str, c["restaurant_norm"], c["restaurant_display"], c["city"],
                  c["stare_akcije"], c["nove_akcije"]) for c in changes]
            )
    except Exception as e:
        import logging as _log
        _log.getLogger("scheduler").error(f"save_promo_history greška: {e}")

def load_promo_history(days: int = 30) -> pd.DataFrame:
    try:
        cutoff = (datetime.datetime.now() - datetime.timedelta(days=days)).strftime("%Y-%m-%d %H:%M:%S")
        with sqlite3.connect(DB_FILE, check_same_thread=False) as conn:
            return pd.read_sql_query(
                "SELECT * FROM promo_history WHERE timestamp >= ? ORDER BY timestamp DESC",
                conn, params=(cutoff,)
            )
    except Exception:
        return pd.DataFrame()

def should_send_am_alert(restaurant_norm: str, city: str, new_akcije_filtered: str, state: dict) -> bool:
    """
    Pametna logika slanja AM alerta:
    - Nije imao akciju → sad ima            → POŠALJI
    - Imao → ista akcija                    → NE ŠALJI
    - Imao → drugačija akcija               → POŠALJI
    - Imao → nema → sad ponovo ima          → POŠALJI (jer je u state-u poslednje stanje '-')
    """
    if new_akcije_filtered == "-":
        return False
    prev = state.get((restaurant_norm, city))
    if prev is None:
        return True   # prvi put vidimo ovaj restoran
    if prev == "-":
        return True   # imao → nema → ponovo ima
    # normalizuj redosled linija da poredimo sadržaj, ne redosled
    def norm_set(s):
        return frozenset(l.strip().lower() for l in s.split("\n") if l.strip())
    return norm_set(prev) != norm_set(new_akcije_filtered)

# ── Aliasi ────────────────────────────────────────────────────────────────────
def load_amm() -> pd.DataFrame:        return load_amm_gsheet()
def save_amm(df):                       save_amm_gsheet(df)
def load_alert_log() -> pd.DataFrame:  return load_alert_log_gsheet()
def append_alert_log(rows):             append_alert_log_gsheet(rows)
def save_scan(df):                      save_scan_gsheet(df)
def load_scan() -> pd.DataFrame:       return load_scan_gsheet()
def scan_meta() -> str | None:         return scan_meta_gsheet()
def load_sales() -> dict:              return load_sales_gsheet()
def save_sales(data):                  save_sales_gsheet(data)

# ─────────────────────────── SCAN LOCK (SQLite) ──────────────────────────────
# Atomičan lock u SQLite — nema zaostalih fajlova ako proces padne.

_LOCK_TIMEOUT_SEC = 10800  # 3h

def _ensure_lock_table():
    with sqlite3.connect(DB_FILE, check_same_thread=False) as conn:
        conn.execute("""CREATE TABLE IF NOT EXISTS scan_lock (
            id          INTEGER PRIMARY KEY CHECK (id = 1),
            locked_at   REAL
        )""")

_ensure_lock_table()

def acquire_scan_lock() -> bool:
    try:
        now = time.time()
        with sqlite3.connect(DB_FILE, check_same_thread=False) as conn:
            conn.execute(
                "DELETE FROM scan_lock WHERE id=1 AND ? - locked_at > ?",
                (now, _LOCK_TIMEOUT_SEC)
            )
            try:
                conn.execute("INSERT INTO scan_lock (id, locked_at) VALUES (1, ?)", (now,))
                return True
            except sqlite3.IntegrityError:
                return False
    except Exception:
        return False

def release_scan_lock():
    try:
        with sqlite3.connect(DB_FILE, check_same_thread=False) as conn:
            conn.execute("DELETE FROM scan_lock WHERE id=1")
    except Exception:
        pass

def is_scan_locked() -> bool:
    try:
        now = time.time()
        with sqlite3.connect(DB_FILE, check_same_thread=False) as conn:
            row = conn.execute("SELECT locked_at FROM scan_lock WHERE id=1").fetchone()
        if row is None:
            return False
        return (now - row[0]) < _LOCK_TIMEOUT_SEC
    except Exception:
        return False

# ─────────────────────────── DEDUPLICATION NOVIH RESTORANA (SQLite) ─────────

def load_sent_new_restaurants() -> set:
    try:
        with sqlite3.connect(DB_FILE, check_same_thread=False) as conn:
            rows = conn.execute("SELECT slug FROM sent_new_restaurants").fetchall()
            return {r[0] for r in rows}
    except Exception:
        return set()

def save_sent_new_restaurants(slugs: set):
    try:
        with sqlite3.connect(DB_FILE, check_same_thread=False) as conn:
            conn.executemany(
                "INSERT OR IGNORE INTO sent_new_restaurants (slug) VALUES (?)",
                [(s,) for s in slugs]
            )
    except Exception:
        pass

# ─────────────────────────── WATCHLIST (SQLite) ───────────────────────────────

def load_watchlist() -> pd.DataFrame:
    try:
        with sqlite3.connect(DB_FILE, check_same_thread=False) as conn:
            return pd.read_sql_query("SELECT * FROM watchlist ORDER BY added_at DESC", conn)
    except Exception:
        return pd.DataFrame(columns=["slug","city","naziv","added_at"])

def add_to_watchlist(slug: str, city: str, naziv: str):
    try:
        with sqlite3.connect(DB_FILE, check_same_thread=False) as conn:
            conn.execute(
                "INSERT OR REPLACE INTO watchlist (slug, city, naziv, added_at) VALUES (?,?,?,?)",
                (slug, city, naziv, local_now())
            )
    except Exception:
        pass

def remove_from_watchlist(slug: str, city: str):
    try:
        with sqlite3.connect(DB_FILE, check_same_thread=False) as conn:
            conn.execute("DELETE FROM watchlist WHERE slug=? AND city=?", (slug, city))
    except Exception:
        pass

def get_watchlist_alerts(df_scan: pd.DataFrame) -> pd.DataFrame:
    """Vraća redove iz skena koji su na watchlisti sa njihovim trenutnim akcijama."""
    wl = load_watchlist()
    if wl.empty or df_scan.empty or "slug" not in df_scan.columns:
        return pd.DataFrame()
    merged = wl.merge(df_scan[["slug","grad","naziv","akcije","status","link"]],
                      on="slug", how="left", suffixes=("_wl",""))
    return merged



def load_alert_cooldown() -> dict:
    try:
        with sqlite3.connect(DB_FILE, check_same_thread=False) as conn:
            rows = conn.execute("SELECT am_email, restaurant_norm, last_sent_date FROM alert_cooldown").fetchall()
            return {f"{r[0]}|{r[1]}": r[2] for r in rows}
    except Exception:
        return {}

def save_alert_cooldown(cooldown: dict):
    try:
        with sqlite3.connect(DB_FILE, check_same_thread=False) as conn:
            conn.executemany(
                "INSERT INTO alert_cooldown (am_email, restaurant_norm, last_sent_date) VALUES (?,?,?) "
                "ON CONFLICT(am_email, restaurant_norm) DO UPDATE SET last_sent_date=excluded.last_sent_date",
                [(k.split("|")[0], k.split("|")[1], v) for k, v in cooldown.items() if "|" in k]
            )
    except Exception:
        pass

def is_in_cooldown(am_email: str, restaurant_norm: str, cooldown: dict) -> bool:
    key = f"{am_email}|{restaurant_norm}"
    last_sent_str = cooldown.get(key)
    if not last_sent_str:
        return False
    try:
        last_sent = datetime.date.fromisoformat(last_sent_str)
        return (datetime.date.today() - last_sent).days < COOLDOWN_DAYS
    except Exception:
        return False

def update_cooldown(am_email: str, restaurant_norm: str, cooldown: dict):
    key = f"{am_email}|{restaurant_norm}"
    cooldown[key] = datetime.date.today().isoformat()

# ─────────────────────────── WOLT API & SESSION ──────────────────────────────

WOLT_COOKIE = ""

BROWSER_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Accept": "application/json, text/plain, */*",
    "Accept-Language": "sr;q=0.9,en-US;q=0.8,en;q=0.7",
    "Origin": "https://wolt.com",
    "Referer": "https://wolt.com/en/srb/",
    "W-PlatformType": "Web",
    "W-Wolt-Session-Id": "wolt-monitor-session",
}

session = requests.Session()
session.headers.update(BROWSER_HEADERS)

# ─── Wolt session auto-refresh ───────────────────────────────────────────────
# Kada API vrati 401/403, skripta automatski pokušava da obnovi session
# kroz anonimni init poziv prema Wolt API-ju.

_session_lock = threading.Lock()
_last_refresh_time = 0.0

def _refresh_wolt_session() -> bool:
    """
    Pokušava da obnovi Wolt session bez cookie-ja.
    Wolt dozvoljava guest browsing — init poziv dobija validan session.
    """
    global _last_refresh_time
    with _session_lock:
        # Anti-throttle: ne refreshuj češće od jednom u 60s
        now = time.time()
        if now - _last_refresh_time < 60:
            return True  # pretpostavi da je refresh već urađen
        try:
            # Koristimo inicijalni API poziv koji ne zahteva auth
            init_url = "https://restaurant-api.wolt.com/v1/pages/restaurants?lat=44.8178&lon=20.4569&skip=0"
            r = requests.get(init_url, headers=BROWSER_HEADERS, timeout=15)
            if r.status_code == 200:
                # Preuzmi kolačiće iz odgovora i ažuriraj globalnu sesiju
                session.cookies.update(r.cookies)
                _last_refresh_time = now
                import logging as _log
                _log.getLogger("scheduler").info("[WoltSession] Session refreshovan uspešno.")
                return True
        except Exception as e:
            import logging as _log
            _log.getLogger("scheduler").warning(f"[WoltSession] Refresh greška: {e}")
        return False

CITY_SLUG_MAP = {
    "Beograd":     "belgrade",
    "Novi Sad":    "novi-sad",
    "Nis":         "nis",
    "Kragujevac":  "kragujevac",
    "Arandelovac": "arandelovac",
    "Bor":         "bor",
    "Borca":       "borca",
    "Cacak":       "cacak",
    "Jagodina":    "jagodina",
    "Kraljevo":    "kraljevo",
    "Krusevac":    "krusevac",
    "Lazarevac":   "lazarevac",
    "Leskovac":    "leskovac",
    "Novi Pazar":  "novi-pazar",
    "Obrenovac":   "obrenovac",
    "Pancevo":     "pancevo",
    "Pozarevac":   "pozarevac",
    "Smederevo":   "smederevo",
    "Sombor":      "sombor",
    "Subotica":    "subotica",
    "Uzice":       "uzice",
    "Valjevo":     "valjevo",
    "Vrsac":       "vrsac",
    "Zlatibor":    "zlatibor",
    "Zrenjanin":   "zrenjanin",
}

def wolt_get(url: str) -> tuple:
    try:
        with _global_http_sem:
            r = session.get(url, timeout=15)
        if r.status_code == 200:
            return r.json(), 200
        if r.status_code in (401, 403):
            _refresh_wolt_session()
            with _global_http_sem:
                r2 = session.get(url, timeout=15)
            if r2.status_code == 200:
                return r2.json(), 200
            return None, r2.status_code
        return None, r.status_code
    except Exception:
        return None, -1

def make_thread_session() -> requests.Session:
    s = requests.Session()
    for k, v in session.headers.items():
        s.headers[k] = v
    try:
        cookie_val = Path("_scan_cookie.txt").read_text().strip()
    except Exception:
        cookie_val = ""
    if not cookie_val:
        cookie_val = WOLT_COOKIE or ""
    if cookie_val:
        s.headers["Cookie"] = cookie_val
    return s

# ─────────────────────────── FETCH AKCIJA ────────────────────────────────────

_fetch_log_lock = threading.Lock()
_throttle_until = 0.0
_throttle_lock  = threading.Lock()

def _log_fetch(msg: str):
    try:
        ts = datetime.datetime.now().strftime("%H:%M:%S")
        with _fetch_log_lock:
            with open("_fetch_debug.log", "a", encoding="utf-8") as f:
                f.write(f"[{ts}] {msg}\n")
    except Exception:
        pass

def _wait_throttle():
    now = time.time()
    with _throttle_lock:
        wait = _throttle_until - now
    if wait > 0:
        time.sleep(wait)

def _set_throttle(seconds: float):
    with _throttle_lock:
        global _throttle_until
        _throttle_until = max(_throttle_until, time.time() + seconds)

def _fetch_url(ts, url: str, label: str, stop_event) -> tuple:
    for attempt in range(4):
        if stop_event.is_set():
            return None, 0
        _wait_throttle()
        try:
            time.sleep(random.uniform(0.3, 1.2))
            with _global_http_sem:          # globalno ograničenje — sve HTTP ka Woltu
                r = ts.get(url, timeout=10)
            if r.status_code == 200:
                return r.json(), 200
            if r.status_code in (401, 403):
                    _log_fetch(f"{label} → {r.status_code} (auth fail) — pokušavam session refresh")
                    _refresh_wolt_session()
                    return None, r.status_code
            if r.status_code == 429:
                wait = 2 + 2 ** attempt
                _set_throttle(wait)
                _log_fetch(f"{label} → 429 retry {attempt} (throttle {wait:.0f}s)")
                continue
            _log_fetch(f"{label} → {r.status_code}")
            return None, r.status_code
        except Exception as e:
            _log_fetch(f"{label} → EXC {e}")
            if attempt < 3:
                time.sleep(0.5)
    return None, -1

def _fetch_one(slug: str, lat: float, lon: float, feed_akcije: list, stop_event: threading.Event) -> tuple[str, str]:
    if stop_event.is_set():
        return slug, "-"
    ts = make_thread_session()
    time.sleep(random.uniform(1.0, 2.0))
    dyn_url = (
        f"https://consumer-api.wolt.com/order-xp/web/v1/venue/slug/{slug}/dynamic/"
        f"?lat={lat}&lon={lon}&selected_delivery_method=homedelivery"
    )
    akcije_str = "-"
    dyn_data, _ = _fetch_url(ts, dyn_url, f"DYN {slug}", stop_event)
    if dyn_data:
        try:
            parsed   = _parse_dynamic_with_item_discount(dyn_data)
            combined = list(dict.fromkeys(feed_akcije + parsed))
            akcije_str = "\n".join(combined) if combined else "-"
            if akcije_str == "-":
                _log_fetch(f"DYN {slug} → 200 ali NEMA akcija")
        except Exception as e:
            _log_fetch(f"DYN {slug} → parse EXC {e}")
    elif feed_akcije:
        akcije_str = "\n".join(feed_akcije)
    return slug, akcije_str

def _parse_dynamic_with_item_discount(data: dict) -> list:
    akcije = []
    seen = set()
    ignore_texts = {
        "prikaži detalje", "show details", "vidi sve", "see all",
        "detalji restorana", "restaurant details", "more", "još",
        "schedule order", "naruči", "see menu", "add {amount} more",
        "try for 30 days for free!", "get rsd0 delivery fee & more!",
    }

    def add(text, wolt_plus=False):
        t = (text or "").strip()
        if not t or len(t) <= 3 or t.lower() in ignore_texts:
            return
        prefix = "• [Wolt+] " if wolt_plus else "• "
        key = t.lower()
        if key not in seen:
            seen.add(key)
            akcije.append(f"{prefix}{t}")

    venue_raw = data.get("venue_raw") or {}
    for disc in venue_raw.get("discounts", []):
        if not isinstance(disc, dict):
            continue
        is_wp = (disc.get("has_wolt_plus") or
                 (disc.get("banner") or {}).get("show_wolt_plus", False) or
                 (disc.get("conditions") or {}).get("has_wolt_plus") == True)
        banner = disc.get("banner") or {}
        desc   = disc.get("description") or {}
        primary_text = banner.get("formatted_text") or desc.get("title") or ""
        add(primary_text, wolt_plus=is_wp)
        effects = disc.get("effects") or {}
        item_discount_dict = effects.get("item_discount")
        if item_discount_dict and isinstance(item_discount_dict, dict):
            fraction = item_discount_dict.get("fraction")
            if fraction and float(fraction) > 0:
                pct = int(round(float(fraction) * 100))
                fallback = primary_text or f"{pct}% popust na izabrane artikle"
                add(fallback, wolt_plus=is_wp)
        basket_disc = effects.get("basket_discount")
        if basket_disc and isinstance(basket_disc, dict):
            amount   = basket_disc.get("amount")
            fraction = basket_disc.get("fraction")
            if amount and int(amount) > 0:
                rsd = int(amount) // 100
                fallback = primary_text or f"{rsd} RSD popust na korpu"
                add(fallback, wolt_plus=is_wp)
            elif fraction and float(fraction) > 0:
                pct = int(round(float(fraction) * 100))
                fallback = primary_text or f"{pct}% popust na celu korpu"
                add(fallback, wolt_plus=is_wp)
        delivery_disc = effects.get("delivery_discount")
        if delivery_disc and isinstance(delivery_disc, dict):
            amount   = delivery_disc.get("amount")
            fraction = delivery_disc.get("fraction")
            if (amount is not None and int(amount) == 0) or (fraction and float(fraction) >= 1.0):
                fallback = primary_text or "Besplatna dostava"
                add(fallback, wolt_plus=is_wp)
            elif amount and int(amount) > 0:
                rsd = int(amount) // 100
                fallback = primary_text or f"{rsd} RSD popust na dostavu"
                add(fallback, wolt_plus=is_wp)
        free_items = effects.get("free_items")
        if free_items and isinstance(free_items, (dict, list)):
            fallback = primary_text or "Gratis artikal uz porudžbinu"
            add(fallback, wolt_plus=is_wp)

    venue = data.get("venue") or {}
    for ban in venue.get("banners", []):
        if not isinstance(ban, dict):
            continue
        is_wp = ban.get("show_wolt_plus", False)
        disc = ban.get("discount") or {}
        add(disc.get("formatted_text"), wolt_plus=is_wp)

    offer_assistant = venue.get("offer_assistant") or {}
    for tracker in offer_assistant.get("offer_trackers", []):
        if not isinstance(tracker, dict):
            continue
        is_wp = tracker.get("offer_type") == "wolt_plus" or tracker.get("show_wolt_plus", False)
        add(tracker.get("title"), wolt_plus=is_wp)

    return akcije


# ─────────────────────────── FETCH GRAD ──────────────────────────────────────

_city_progress = {}
_city_progress_lock = threading.Lock()

def _update_city_progress(city_display: str, found: int = None, total: int = None, status: str = None):
    with _city_progress_lock:
        if city_display not in _city_progress:
            _city_progress[city_display] = {"found": 0, "total": 0, "status": "čekanje..."}
        if found is not None:
            _city_progress[city_display]["found"] = max(_city_progress[city_display]["found"], found)
        if total is not None:
            _city_progress[city_display]["total"] = max(_city_progress[city_display]["total"], total)
        if status is not None:
            _city_progress[city_display]["status"] = status

def _write_status_file():
    with _city_progress_lock:
        data = dict(_city_progress)
    try:
        tmp = Path("_scan_city_progress.json.tmp")
        tmp.write_text(json.dumps(data, ensure_ascii=False), encoding="utf-8")
        tmp.replace(Path("_scan_city_progress.json"))
    except Exception:
        pass

def fetch_city(city_display: str, status_placeholder, stop_event: threading.Event) -> list[dict]:
    city_key  = display_to_key(city_display)
    city_slug = CITY_SLUG_MAP.get(city_key)
    multi_coords = get_active_coords().get(city_key, CITY_MULTI_COORDS.get(city_key, [CITY_COORDS.get(city_key, (44.8178, 20.4569))]))
    primary_lat, primary_lon = multi_coords[0]

    if not city_slug:
        status_placeholder.error(f"❌ Nepoznat grad: '{city_display}'")
        return []

    restaurants = {}
    _update_city_progress(city_display, found=0, total=0, status="Loading restaurant list...")
    _write_status_file()

    for loc_idx, (lat, lon) in enumerate(multi_coords):
        if stop_event.is_set():
            break
        loc_label = f"lok. {loc_idx+1}/{len(multi_coords)}"
        skip = 0
        for page_num in range(50):
            if stop_event.is_set():
                break
            count_before = len(restaurants)
            endpoint = f"https://restaurant-api.wolt.com/v1/pages/restaurants?lat={lat}&lon={lon}&skip={skip}"
            data, _status = wolt_get(endpoint)
            items_in_response = 0
            if data:
                for section in data.get("sections", []):
                    for item in section.get("items", []):
                        venue = item.get("venue")
                        if not venue:
                            continue
                        name = venue.get("name", "")
                        slug = venue.get("slug", "")
                        if not name or not slug or slug in restaurants:
                            continue
                        items_in_response += 1
                        status_obj = "Otvoren" if venue.get("online") else "Zatvoren"
                        rating   = venue.get("rating") or {}
                        r_score  = rating.get("score", "-") if isinstance(rating, dict) else "-"
                        est      = venue.get("estimate_range") or venue.get("estimate")
                        delivery = f"{est} min" if est else "-"
                        feed_akcije = []
                        novo_status = "Ne"
                        for badge in venue.get("badges", []):
                            txt = badge.get("text", "")
                            if txt:
                                if txt.lower() in ["novo", "new"]:
                                    novo_status = "Da"
                                else:
                                    feed_akcije.append(f"• {txt}")
                        label = venue.get("label", "")
                        if label:
                            if label.lower() in ["novo", "new"]:
                                novo_status = "Da"
                            else:
                                feed_akcije.append(f"• {label}")
                        restaurants[slug] = {
                            "grad":         city_display,
                            "naziv":        name,
                            "slug":         slug,
                            "status":       status_obj,
                            "ocena":        str(r_score),
                            "dostava":      delivery,
                            "novo":         novo_status,
                            "_feed_akcije": feed_akcije,
                            "akcije":       "-",
                            "link":         f"https://wolt.com/en/srb/{city_slug}/restaurant/{slug}",
                            "naziv_norm":   normalize(name),
                        }
            new_this_page = len(restaurants) - count_before
            _update_city_progress(city_display, found=len(restaurants),
                                  status=f"📍 {loc_label} | str.{page_num+1} +{new_this_page} (total {len(restaurants)})")
            _write_status_file()
            if items_in_response == 0:
                break
            skip += 40
            time.sleep(random.uniform(0.5, 1.8))
        _update_city_progress(city_display, status=f"✅ Location {loc_idx+1}/{len(multi_coords)} done ({len(restaurants)} total)")
        _write_status_file()

    if not restaurants or stop_event.is_set():
        if not restaurants:
            _update_city_progress(city_display, status="⚠️ No restaurants found.")
        _write_status_file()
        return []

    slugs = list(restaurants.keys())
    total = len(slugs)
    completed = 0
    _update_city_progress(city_display, total=total, found=total,
                          status=f"⚡ Loading promotions (0/{total})...")
    _write_status_file()

    with ThreadPoolExecutor(max_workers=FETCH_WORKERS) as executor:
        futures = {
            executor.submit(
                _fetch_one, slug, primary_lat, primary_lon,
                restaurants[slug]["_feed_akcije"], stop_event,
            ): slug for slug in slugs
        }
        for future in as_completed(futures):
            if stop_event.is_set():
                executor.shutdown(wait=False, cancel_futures=True)
                break
            try:
                slug, akcije_str = future.result()
                restaurants[slug]["akcije"] = akcije_str
            except Exception:
                pass
            completed += 1
            if completed % 10 == 0 or completed == total:
                _update_city_progress(city_display, status=f"⚡ Promotions: {completed}/{total} restaurants")
                _write_status_file()

    for r in restaurants.values():
        r.pop("_feed_akcije", None)

    _update_city_progress(city_display, status=f"✅ Done! {len(restaurants)} restaurants")
    _write_status_file()
    return list(restaurants.values())

def scan_all_cities(selected_cities: list[str], status_placeholder, stop_event: threading.Event) -> pd.DataFrame:
    with _city_progress_lock:
        _city_progress.clear()
    for city in selected_cities:
        _update_city_progress(city, found=0, total=0, status="⏳ Waiting in queue...")
    _write_status_file()

    all_rows = []
    all_rows_lock = threading.Lock()

    def _fetch_city_safe(city: str):
        if stop_event.is_set():
            return
        try:
            rows = fetch_city(city, status_placeholder, stop_event)
            with all_rows_lock:
                all_rows.extend(rows)
        except Exception as e:
            _update_city_progress(city, status=f"❌ Error: {e}")
            _write_status_file()

    # Gradovi idu paralelno (max CITY_PARALLEL odjednom), ali ukupan broj
    # HTTP zahteva ka Woltu ostaje ograničen globalnim semaforem _global_http_sem.
    with ThreadPoolExecutor(max_workers=CITY_PARALLEL) as city_executor:
        futures = [city_executor.submit(_fetch_city_safe, city) for city in selected_cities]
        for future in as_completed(futures):
            if stop_event.is_set():
                city_executor.shutdown(wait=False, cancel_futures=True)
                break
            try:
                future.result()
            except Exception:
                pass

    status_placeholder.empty()
    return pd.DataFrame(all_rows) if all_rows else pd.DataFrame()

def scan_nopromo_cities(selected_cities: list[str], prev_df: pd.DataFrame, stop_event: threading.Event) -> pd.DataFrame:
    with _city_progress_lock:
        _city_progress.clear()
    no_promo = prev_df[
        (prev_df["grad"].isin(selected_cities)) & (prev_df["akcije"] == "-")
    ].copy()
    other = prev_df[~prev_df["grad"].isin(selected_cities)].copy()
    had_promo = prev_df[
        (prev_df["grad"].isin(selected_cities)) & (prev_df["akcije"] != "-")
    ].copy()

    for city in selected_cities:
        city_count = len(no_promo[no_promo["grad"] == city])
        _update_city_progress(city, found=city_count, total=city_count,
                              status=f"⏳ Waiting in queue... ({city_count} restorana za sken)")
    _write_status_file()

    updated_rows = []
    for city in selected_cities:
        if stop_event.is_set():
            break
        city_key = display_to_key(city)
        all_coords = get_active_coords().get(city_key, CITY_MULTI_COORDS.get(city_key, [(44.8178, 20.4569)]))
        # Koristimo srednju koordinatu kao reprezentativnu tačku za No-Promo sken
        primary_lat = sum(c[0] for c in all_coords) / len(all_coords)
        primary_lon = sum(c[1] for c in all_coords) / len(all_coords)
        city_subset = no_promo[no_promo["grad"] == city]
        slugs = city_subset["slug"].tolist() if "slug" in city_subset.columns else []
        total = len(slugs)
        completed = 0
        _update_city_progress(city, found=total, total=total, status=f"⚡ Scanning promotions (0/{total})...")
        _write_status_file()
        slug_to_row = {row["slug"]: row.to_dict() for _, row in city_subset.iterrows()} if "slug" in city_subset.columns else {}

        with ThreadPoolExecutor(max_workers=FETCH_WORKERS) as executor:
            futures = {
                executor.submit(_fetch_one, slug, primary_lat, primary_lon, [], stop_event): slug
                for slug in slugs
            }
            for future in as_completed(futures):
                if stop_event.is_set():
                    executor.shutdown(wait=False, cancel_futures=True)
                    break
                try:
                    slug, akcije_str = future.result()
                    row = dict(slug_to_row.get(slug, {}))
                    row["akcije"] = akcije_str
                    updated_rows.append(row)
                except Exception:
                    pass
                completed += 1
                if completed % 10 == 0 or completed == total:
                    _update_city_progress(city, status=f"⚡ Akcije: {completed}/{total}")
                    _write_status_file()
        _update_city_progress(city, status=f"✅ Done! {total} restaurants scanned")
        _write_status_file()

    all_parts = []
    if updated_rows:
        all_parts.append(pd.DataFrame(updated_rows))
    if not had_promo.empty:
        all_parts.append(had_promo)
    if not other.empty:
        all_parts.append(other)
    if all_parts:
        return pd.concat(all_parts, ignore_index=True)
    return prev_df.copy()

# ─────────────────────────── EMAIL ───────────────────────────────────────────

def send_alert_email(am_email: str, am_name: str, alerts: list[dict], all_partners: list[dict] = None, tip: str = "AM") -> bool:
    try:
        # ── Gornji deo: novi/promenjeni alertovi ─────────────────────────────
        rows_html = ""
        for a in alerts:
            akcije_filtered = filter_akcije_for_email(a["akcije"])
            if akcije_filtered != "-":
                akcije_html = akcije_filtered.replace("\n", "<br>")
            else:
                akcije_html = "<span style='color:#aaa'>–</span>"
            link = a.get("link", "")
            if link:
                naziv_cell = f"<a href='{link}' style='color:#222;text-decoration:none;font-weight:600'>{a['naziv']}</a>"
            else:
                naziv_cell = f"<span style='font-weight:600'>{a['naziv']}</span>"
            rows_html += f"""
            <tr>
              <td style='padding:10px 14px;border-bottom:1px solid #eee'>{naziv_cell}</td>
              <td style='padding:10px 14px;border-bottom:1px solid #eee;color:#555'>{a['grad']}</td>
              <td style='padding:10px 14px;border-bottom:1px solid #eee;color:#333'>{akcije_html}</td>
            </tr>"""
        if not rows_html:
            return True

        # ── Donji deo: svi partneri tog AM-a iz poslednjeg skena ─────────────
        all_partners_html = ""
        if all_partners:
            # Prikazuj SAMO partnere koji imaju aktivne promocije
            partners_sa_promo = [p for p in all_partners if filter_akcije_for_email(p.get("akcije","")) != "-"]
            sa_promo = len(partners_sa_promo)
            ukupno   = len(all_partners)

            sorted_partners = sorted(partners_sa_promo, key=lambda x: x.get("naziv",""))
            for p in sorted_partners:
                akcije_filtered = filter_akcije_for_email(p.get("akcije", ""))
                akcije_html = akcije_filtered.replace("\n", "<br>")
                dot = "<span style='color:#27ae60;font-weight:700'>●</span>"
                link = p.get("link", "")
                if link:
                    naziv_cell = f"<a href='{link}' style='color:#222;text-decoration:none;font-weight:600'>{p['naziv']}</a>"
                else:
                    naziv_cell = f"<span style='font-weight:600'>{p['naziv']}</span>"
                all_partners_html += f"""
                <tr>
                  <td style='padding:8px 14px;border-bottom:1px solid #f0f0f0;width:20px'>{dot}</td>
                  <td style='padding:8px 14px;border-bottom:1px solid #f0f0f0'>{naziv_cell}</td>
                  <td style='padding:8px 14px;border-bottom:1px solid #f0f0f0;color:#555'>{p.get('grad','')}</td>
                  <td style='padding:8px 14px;border-bottom:1px solid #f0f0f0;color:#333;font-size:13px'>{akcije_html}</td>
                </tr>"""

            all_partners_section = f"""
            <div style='margin-top:32px;border-top:2px solid #eee;padding-top:24px'>
              <h3 style='color:#1a1a2e;margin:0 0 14px 0;font-size:1rem'>
                📋 Trenutno stanje tvojih partnera sa promocijama
              </h3>
              <div style='display:inline-flex;gap:24px;margin-bottom:18px;align-items:center'>
                <div style='text-align:center;background:#f0f4ff;border-radius:10px;padding:10px 22px'>
                  <div style='font-size:2rem;font-weight:800;color:#1a1a2e;line-height:1'>{ukupno}</div>
                  <div style='font-size:11px;color:#888;margin-top:4px;letter-spacing:0.5px'>PARTNERA UKUPNO</div>
                </div>
                <div style='text-align:center;background:#eafaf1;border-radius:10px;padding:10px 22px;border:2px solid #27ae60'>
                  <div style='font-size:2rem;font-weight:800;color:#27ae60;line-height:1'>{sa_promo}</div>
                  <div style='font-size:11px;color:#27ae60;margin-top:4px;letter-spacing:0.5px'>AKTIVNIH PROMOCIJA</div>
                </div>
              </div>
              <table style='border-collapse:collapse;width:100%;font-size:14px'>
                <thead>
                  <tr style='background:#f7f8fc'>
                    <th style='padding:8px 14px;text-align:left;color:#1a1a2e;border-bottom:2px solid #eee;width:20px'></th>
                    <th style='padding:8px 14px;text-align:left;color:#1a1a2e;border-bottom:2px solid #eee'>Restoran</th>
                    <th style='padding:8px 14px;text-align:left;color:#1a1a2e;border-bottom:2px solid #eee'>Grad</th>
                    <th style='padding:8px 14px;text-align:left;color:#1a1a2e;border-bottom:2px solid #eee'>Akcije</th>
                  </tr>
                </thead>
                <tbody>{all_partners_html}</tbody>
              </table>
            </div>"""
        else:
            all_partners_section = ""

        today_str = datetime.date.today().strftime("%d.%m.%Y")
        html = f"""
        <html><body style='font-family:Arial,sans-serif;color:#222;max-width:720px;margin:auto'>
          <div style='background:#1a1a2e;padding:24px 32px;border-radius:12px 12px 0 0'>
            <h2 style='color:#fff;margin:0'>📊 Promo Monitor – {today_str}</h2>
          </div>
          <div style='background:#fff;padding:24px 32px;border-radius:0 0 12px 12px;
                      box-shadow:0 4px 16px rgba(0,0,0,0.08)'>
            <p>Zdravo <b>{am_name}</b>,</p>
            <p>Sledeći tvoji partneri imaju <b>nove ili promenjene promotivne akcije na konkurenciji</b>:</p>
            <table style='border-collapse:collapse;width:100%;font-size:14px'>
              <thead>
                <tr style='background:#f0f4ff'>
                  <th style='padding:10px 14px;text-align:left;color:#1a1a2e;border-bottom:2px solid #dde'>Restoran</th>
                  <th style='padding:10px 14px;text-align:left;color:#1a1a2e;border-bottom:2px solid #dde'>Grad</th>
                  <th style='padding:10px 14px;text-align:left;color:#1a1a2e;border-bottom:2px solid #dde'>Akcije</th>
                </tr>
              </thead>
              <tbody>{rows_html}</tbody>
            </table>
            {all_partners_section}
            <p style='margin-top:20px;font-size:12px;color:#999'>
              Automatski izveštaj &bull; {local_now()}
            </p>
          </div>
        </body></html>"""
        msg = MIMEMultipart("alternative")
        msg["From"]    = EMAIL_SENDER
        msg["To"]      = am_email
        msg["Subject"] = f"📊 Promo izveštaj – {len(alerts)} partnera – {today_str}"
        msg.attach(MIMEText(html, "html"))
        with smtplib.SMTP("smtp.gmail.com", 587) as srv:
            srv.starttls()
            srv.login(EMAIL_SENDER, EMAIL_PASSWORD)
            srv.sendmail(EMAIL_SENDER, am_email, msg.as_string())
        # Slack webhook (ako je konfigurisan)
        slack_url = load_slack_webhook()
        if slack_url:
            send_slack_notification(slack_url, alerts, am_name)
        return True
    except Exception as e:
        import logging as _log
        _log.getLogger("scheduler").error(f"Email greška ({am_email}): {e}")
        try:
            st.error(f"Email greška ({am_email}): {e}")
        except Exception:
            pass
        return False

def send_sales_bulk_notification(to_email: str, grad: str, novi_restorani: list) -> bool:
    try:
        rows_html = ""
        for r in novi_restorani:
            naziv = r.get("naziv", "")
            slug  = r.get("slug", "")
            grad_slug = (grad.lower()
                         .replace(" ", "-")
                         .replace("š", "s").replace("Š", "s")
                         .replace("ć", "c").replace("Ć", "c")
                         .replace("č", "c").replace("Č", "c")
                         .replace("đ", "dj").replace("Đ", "dj")
                         .replace("ž", "z").replace("Ž", "z"))
            wolt_link = f"https://wolt.com/sr/srb/{grad_slug}/restaurant/{slug}"
            rows_html += f"""
            <tr>
              <td style='padding:10px 14px;border-bottom:1px solid #eee;font-weight:600'>{naziv}</td>
              <td style='padding:10px 14px;border-bottom:1px solid #eee'>
                <a href='{wolt_link}' style='color:#009de0'>{wolt_link}</a>
              </td>
            </tr>"""

        today_str = datetime.date.today().strftime("%d.%m.%Y")
        html = f"""
        <html><body style='font-family:Arial,sans-serif;background:#f5f5f5;padding:20px'>
          <div style='max-width:680px;margin:auto;background:#fff;border-radius:12px;
                      padding:28px;box-shadow:0 2px 8px rgba(0,0,0,0.08)'>
            <div style='background:#009de0;color:#fff;padding:16px 24px;border-radius:8px;
                        margin-bottom:20px;font-size:1.2rem;font-weight:700'>
              🆕 Novi restorani na Woltu — {grad} — {today_str}
            </div>
            <p>Na konkurenciji su detektovani novi restorani u gradu <b>{grad}</b> koji nemaju dodeljenog Account Managera:</p>
            <table style='border-collapse:collapse;width:100%;font-size:14px'>
              <thead>
                <tr style='background:#f0f4ff'>
                  <th style='padding:10px 14px;text-align:left;color:#1a1a2e;border-bottom:2px solid #dde'>Restoran</th>
                  <th style='padding:10px 14px;text-align:left;color:#1a1a2e;border-bottom:2px solid #dde'>Wolt link</th>
                </tr>
              </thead>
              <tbody>{rows_html}</tbody>
            </table>
            <p style='margin-top:20px;color:#666;font-size:0.9rem'>
              Ukupno novih restorana: <b>{len(novi_restorani)}</b><br>
              Molimo vas kontaktirajte restoran i ponudite saradnju.
            </p>
            <p style='font-size:11px;color:#999;margin-top:20px'>
              Automatski izveštaj &bull; Promo Monitor &bull; {local_now()}
            </p>
          </div>
        </body></html>"""

        msg = MIMEMultipart("alternative")
        msg["From"]    = EMAIL_SENDER
        msg["To"]      = to_email
        msg["Subject"] = f"🆕 {len(novi_restorani)} novi restoran(a) — {grad} — {today_str}"
        msg.attach(MIMEText(html, "html"))

        with smtplib.SMTP("smtp.gmail.com", 587) as srv:
            srv.starttls()
            srv.login(EMAIL_SENDER, EMAIL_PASSWORD)
            srv.sendmail(EMAIL_SENDER, to_email, msg.as_string())
        return True
    except Exception as e:
        _log_fetch(f"SALES BULK MAIL greška → {to_email}: {e}")
        return False

def send_digest_email(am_email: str, am_name: str, all_partners: list[dict]) -> bool:
    """
    Šalje nedeljni digest svim AM-ovima — pregled SVIH njihovih partnera,
    bez obzira na promene ili cooldown.
    """
    try:
        partners_sa_promo = [p for p in all_partners if filter_akcije_for_email(p.get("akcije","")) != "-"]
        partners_bez = [p for p in all_partners if filter_akcije_for_email(p.get("akcije","")) == "-"]
        sa_promo_count = len(partners_sa_promo)
        ukupno = len(all_partners)

        def partner_row(p, has_promo=True):
            akcije_filtered = filter_akcije_for_email(p.get("akcije",""))
            akcije_html = akcije_filtered.replace("\n","<br>") if has_promo else "<span style='color:#aaa'>Nema akcije</span>"
            dot = "<span style='color:#27ae60;font-weight:700'>●</span>" if has_promo else "<span style='color:#ccc'>○</span>"
            link = p.get("link","")
            naziv_cell = (f"<a href='{link}' style='color:#222;text-decoration:none;font-weight:600'>{p['naziv']}</a>"
                          if link else f"<span style='font-weight:600'>{p['naziv']}</span>")
            return f"""
            <tr>
              <td style='padding:8px 14px;border-bottom:1px solid #f0f0f0;width:20px'>{dot}</td>
              <td style='padding:8px 14px;border-bottom:1px solid #f0f0f0'>{naziv_cell}</td>
              <td style='padding:8px 14px;border-bottom:1px solid #f0f0f0;color:#555'>{p.get('grad','')}</td>
              <td style='padding:8px 14px;border-bottom:1px solid #f0f0f0;color:#333;font-size:13px'>{akcije_html}</td>
            </tr>"""

        rows_html = "".join(partner_row(p, True) for p in sorted(partners_sa_promo, key=lambda x: x.get("naziv","")))
        rows_html += "".join(partner_row(p, False) for p in sorted(partners_bez, key=lambda x: x.get("naziv","")))

        today_str = datetime.date.today().strftime("%d.%m.%Y")
        html = f"""
        <html><body style='font-family:Arial,sans-serif;color:#222;max-width:720px;margin:auto'>
          <div style='background:#1a1a2e;padding:24px 32px;border-radius:12px 12px 0 0'>
            <h2 style='color:#fff;margin:0'>📋 Nedeljni Digest — {today_str}</h2>
          </div>
          <div style='background:#fff;padding:24px 32px;border-radius:0 0 12px 12px;
                      box-shadow:0 4px 16px rgba(0,0,0,0.08)'>
            <p>Zdravo <b>{am_name}</b>,</p>
            <p>Ovo je tvoj nedeljni pregled stanja promocija za sve tvoje partnere na Woltu.</p>
            <div style='display:inline-flex;gap:24px;margin-bottom:20px;align-items:center'>
              <div style='text-align:center;background:#f0f4ff;border-radius:10px;padding:10px 22px'>
                <div style='font-size:2rem;font-weight:800;color:#1a1a2e;line-height:1'>{ukupno}</div>
                <div style='font-size:11px;color:#888;margin-top:4px;letter-spacing:0.5px'>PARTNERA UKUPNO</div>
              </div>
              <div style='text-align:center;background:#eafaf1;border-radius:10px;padding:10px 22px;border:2px solid #27ae60'>
                <div style='font-size:2rem;font-weight:800;color:#27ae60;line-height:1'>{sa_promo_count}</div>
                <div style='font-size:11px;color:#27ae60;margin-top:4px;letter-spacing:0.5px'>AKTIVNIH PROMOCIJA</div>
              </div>
            </div>
            <table style='border-collapse:collapse;width:100%;font-size:14px'>
              <thead>
                <tr style='background:#f7f8fc'>
                  <th style='padding:8px 14px;text-align:left;color:#1a1a2e;border-bottom:2px solid #eee;width:20px'></th>
                  <th style='padding:8px 14px;text-align:left;color:#1a1a2e;border-bottom:2px solid #eee'>Restoran</th>
                  <th style='padding:8px 14px;text-align:left;color:#1a1a2e;border-bottom:2px solid #eee'>Grad</th>
                  <th style='padding:8px 14px;text-align:left;color:#1a1a2e;border-bottom:2px solid #eee'>Akcije</th>
                </tr>
              </thead>
              <tbody>{rows_html}</tbody>
            </table>
            <p style='margin-top:20px;font-size:12px;color:#999'>
              Nedeljni digest &bull; Promo Monitor &bull; {local_now()}
            </p>
          </div>
        </body></html>"""

        msg = MIMEMultipart("alternative")
        msg["From"]    = EMAIL_SENDER
        msg["To"]      = am_email
        msg["Subject"] = f"📋 Nedeljni digest — {sa_promo_count}/{ukupno} partnera sa promo — {today_str}"
        msg.attach(MIMEText(html, "html"))
        with smtplib.SMTP("smtp.gmail.com", 587) as srv:
            srv.starttls()
            srv.login(EMAIL_SENDER, EMAIL_PASSWORD)
            srv.sendmail(EMAIL_SENDER, am_email, msg.as_string())
        return True
    except Exception as e:
        import logging as _log
        _log.getLogger("scheduler").error(f"Digest email greška ({am_email}): {e}")
        return False



SCHEDULER_FILE = Path("scheduler_config.json")

def load_scheduler_config() -> dict:
    if SCHEDULER_FILE.exists():
        try:
            return json.loads(SCHEDULER_FILE.read_text())
        except Exception:
            pass
    return {"enabled": False, "hour": 8, "minute": 0, "cities": CITIES}

def save_scheduler_config(cfg: dict):
    try:
        SCHEDULER_FILE.write_text(json.dumps(cfg))
    except Exception as e:
        import logging as _log
        _log.getLogger("scheduler").error(f"save_scheduler_config greška: {e}")
        try:
            st.error(f"❌ Greška pri čuvanju scheduler konfiguracije: {e}")
        except Exception:
            pass

def send_slack_notification(webhook_url: str, alerts: list[dict], am_name: str) -> bool:
    """Šalje alert na Slack webhook."""
    if not webhook_url or not alerts:
        return False
    try:
        lines = [f"*📊 Promo Monitor — novi alertovi za {am_name}*"]
        for a in alerts[:10]:
            akcije_filtered = filter_akcije_for_email(a.get("akcije",""))
            akcije_text = akcije_filtered.replace("\n"," | ") if akcije_filtered != "-" else "—"
            lines.append(f"• *{a['naziv']}* ({a.get('grad','')}) — {akcije_text}")
        if len(alerts) > 10:
            lines.append(f"_...i još {len(alerts)-10} partnera_")
        payload = {"text": "\n".join(lines)}
        r = requests.post(webhook_url, json=payload, timeout=10)
        return r.status_code == 200
    except Exception as e:
        import logging as _log
        _log.getLogger("scheduler").error(f"Slack webhook greška: {e}")
        return False

SLACK_WEBHOOK_FILE = Path("slack_webhook.txt")

def load_slack_webhook() -> str:
    try:
        return SLACK_WEBHOOK_FILE.read_text().strip()
    except Exception:
        return ""

def save_slack_webhook(url: str):
    try:
        SLACK_WEBHOOK_FILE.write_text(url.strip())
    except Exception:
        pass

class _NullPH:
    """Dummy Streamlit placeholder za background scheduler pozive."""
    def info(self, *a, **kw): pass
    def warning(self, *a, **kw): pass
    def success(self, *a, **kw): pass
    def error(self, *a, **kw): pass
    def empty(self, *a, **kw): pass

def run_scheduled_scan_and_send():
    import logging
    log = logging.getLogger("scheduler")
    cfg = load_scheduler_config()
    if not cfg.get("enabled"):
        return

    if not acquire_scan_lock():
        log.warning("[Scheduler] Scan već u toku, preskačem zakazani sken.")
        return

    stop_ev = threading.Event()

    try:
        df = scan_all_cities(cfg["cities"], _NullPH(), stop_ev)
    except Exception as e:
        log.error(f"[Scheduler] Scan error: {e}")
        release_scan_lock()
        return

    if df.empty:
        release_scan_lock()
        return

    save_scan(df)
    
    df.to_json("_scan_result.json", orient="records", force_ascii=False)
    Path("_scan_done.txt").write_text("1")

    sent_slugs   = load_sent_new_restaurants()
    amm_df_curr  = load_amm()
    sales_cfg    = load_sales()
    novi_df      = df[df["novo"] == "Da"].copy() if "novo" in df.columns else pd.DataFrame()

    if not novi_df.empty:
        novi_po_gradu  = {}
        new_sent_slugs = set(sent_slugs)
        for _, row in novi_df.iterrows():
            naziv = row.get("naziv", "")
            grad  = row.get("grad", "")
            slug  = row.get("slug", "")
            norm  = normalize(naziv)
            if slug in sent_slugs:
                continue
            has_am = False
            if not amm_df_curr.empty:
                has_am = not amm_df_curr[
                    (amm_df_curr["restaurant_norm"] == norm) & (amm_df_curr["city"] == grad)
                ].empty
            if not has_am:
                if grad not in novi_po_gradu:
                    novi_po_gradu[grad] = []
                novi_po_gradu[grad].append({"naziv": naziv, "slug": slug})
                new_sent_slugs.add(slug)
        for grad, restorani in novi_po_gradu.items():
            for email in sales_cfg.get(grad, []):
                ok = send_sales_bulk_notification(email, grad, restorani)
                if ok:
                    log.info(f"[Scheduler] Bulk sales mail → {email} ({grad}): {len(restorani)} restorana")
        save_sent_new_restaurants(new_sent_slugs)

    if amm_df_curr.empty:
        release_scan_lock()
        return

    df["naziv_norm"] = df["naziv"].apply(normalize)
    merged = df.merge(
        amm_df_curr[["restaurant_norm", "restaurant_display", "city", "am_name", "am_email"]],
        left_on="naziv_norm", right_on="restaurant_norm", how="inner"
    )

    promo_state = load_promo_state()
    state_updates = {}
    sent_log = []
    history_changes = []

    for (am_name, am_email_addr), grp in merged.groupby(["am_name", "am_email"]):
        alerts = []
        for _, row in grp.iterrows():
            rest_norm = normalize(row["naziv"])
            city      = row["grad"]
            akcije_filtered = filter_akcije_for_email(row["akcije"])

            # Uvek ažuriraj state (i za restoran bez akcije → pamtimo '-')
            prev_state = promo_state.get((rest_norm, city))
            state_updates[(rest_norm, city)] = akcije_filtered

            # Snimi promenu u istoriju ako se akcija promenila
            if prev_state is not None and prev_state != akcije_filtered:
                history_changes.append({
                    "restaurant_norm":    rest_norm,
                    "restaurant_display": row["naziv"],
                    "city":               city,
                    "stare_akcije":       prev_state,
                    "nove_akcije":        akcije_filtered,
                })

            if not should_send_am_alert(rest_norm, city, akcije_filtered, promo_state):
                continue
            alerts.append({
                "naziv": row["naziv"],
                "grad":  city,
                "akcije": row["akcije"],
                "link":  row.get("link", ""),
                "norm":  rest_norm,
            })
        if not alerts:
            continue
        all_partners = grp[["naziv", "grad", "akcije", "link"]].rename(
            columns={"naziv": "naziv", "grad": "grad", "akcije": "akcije", "link": "link"}
        ).to_dict("records")
        ok = send_alert_email(am_email_addr, am_name, alerts, all_partners=all_partners, tip="AM-Scheduler")
        if ok:
            for a in alerts:
                sent_log.append({
                    "timestamp":          local_now(),
                    "city":               a["grad"],
                    "restaurant_display": a["naziv"],
                    "am_name":            am_name,
                    "am_email":           am_email_addr,
                    "akcije":             a["akcije"],
                })

    # Sačuvaj novo stanje akcija i istoriju promena
    save_promo_state_bulk(state_updates)
    save_promo_history(history_changes)
    if sent_log:
        append_alert_log(sent_log)

    release_scan_lock()

DAYS_SHORT = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]

def _scheduler_loop():
    import logging
    log = logging.getLogger("scheduler")
    _last_global_fire = None
    _last_city_fire   = {}

    while True:
        try:
            cfg = load_scheduler_config()
            now = datetime.datetime.now()
            today_str = now.strftime("%Y-%m-%d")
            cur_day_short = DAYS_SHORT[now.weekday()]

            # ── Global full scan ─────────────────────────────────────────────
            if cfg.get("enabled"):
                target = now.replace(hour=cfg["hour"], minute=cfg["minute"],
                                     second=0, microsecond=0)
                diff = (now - target).total_seconds()
                fire_key = f"{today_str}_{cfg['hour']:02d}{cfg['minute']:02d}"
                if 0 <= diff < 90 and _last_global_fire != fire_key:
                    _last_global_fire = fire_key
                    log.info(f"[Scheduler] Global scan firing at {now.strftime('%H:%M:%S')}")
                    run_scheduled_scan_and_send()

            # ── Per-city schedules ───────────────────────────────────────────
            city_schedules = cfg.get("city_schedules", {})
            for city, cs in city_schedules.items():
                if not cs.get("enabled") or not cs.get("days"):
                    continue
                if cur_day_short not in cs["days"]:
                    continue
                now2 = datetime.datetime.now()
                target_c = now2.replace(hour=cs["hour"], minute=cs["minute"],
                                        second=0, microsecond=0)
                diff_c = (now2 - target_c).total_seconds()
                fire_key_c = f"{today_str}_{city}_{cs['hour']:02d}{cs['minute']:02d}"
                if 0 <= diff_c < 90 and _last_city_fire.get(city) != fire_key_c:
                    _last_city_fire[city] = fire_key_c
                    log.info(f"[Scheduler-City] {city}: firing at {now2.strftime('%H:%M:%S')}")
                    if not acquire_scan_lock():
                        log.warning(f"[Scheduler-City] {city}: scan already running, skipping.")
                        continue
                    stop_ev = threading.Event()
                    try:
                        prev_df = load_scan()
                        new_city_df = scan_all_cities([city], _NullPH(), stop_ev)
                        if new_city_df is not None and not new_city_df.empty:
                            if not prev_df.empty:
                                other = prev_df[prev_df["grad"] != city]
                                merged_df = pd.concat([other, new_city_df], ignore_index=True)
                            else:
                                merged_df = new_city_df
                            save_scan(merged_df)
                            merged_df.to_json("_scan_result.json", orient="records", force_ascii=False)
                            Path("_scan_done.txt").write_text("1")
                            log.info(f"[Scheduler-City] {city}: done, {len(new_city_df)} restaurants.")

                            # ── Šalji AM alertove za ovaj grad ───────────────
                            amm_df_c = load_amm()
                            if not amm_df_c.empty:
                                new_city_df["naziv_norm"] = new_city_df["naziv"].apply(normalize)
                                merged_am = new_city_df.merge(
                                    amm_df_c[["restaurant_norm", "restaurant_display", "city", "am_name", "am_email"]],
                                    left_on="naziv_norm", right_on="restaurant_norm", how="inner"
                                )
                                if not merged_am.empty:
                                    promo_state = load_promo_state()
                                    state_updates = {}
                                    sent_log = []
                                    history_changes = []
                                    for (am_name_c, am_email_c), grp_c in merged_am.groupby(["am_name", "am_email"]):
                                        alerts_c = []
                                        for _, row_c in grp_c.iterrows():
                                            rn = normalize(row_c["naziv"])
                                            ct = row_c["grad"]
                                            af = filter_akcije_for_email(row_c["akcije"])
                                            prev_s = promo_state.get((rn, ct))
                                            state_updates[(rn, ct)] = af
                                            if prev_s is not None and prev_s != af:
                                                history_changes.append({"restaurant_norm": rn,
                                                    "restaurant_display": row_c["naziv"], "city": ct,
                                                    "stare_akcije": prev_s, "nove_akcije": af})
                                            if not should_send_am_alert(rn, ct, af, promo_state):
                                                continue
                                            alerts_c.append({"naziv": row_c["naziv"], "grad": ct,
                                                             "akcije": row_c["akcije"], "link": row_c.get("link", ""), "norm": rn})
                                        if not alerts_c:
                                            continue
                                        all_p = grp_c[["naziv", "grad", "akcije", "link"]].to_dict("records")
                                        ok_c = send_alert_email(am_email_c, am_name_c, alerts_c, all_partners=all_p, tip="AM-City-Scheduler")
                                        if ok_c:
                                            for ac in alerts_c:
                                                sent_log.append({"timestamp": local_now(), "city": ac["grad"],
                                                    "restaurant_display": ac["naziv"], "am_name": am_name_c,
                                                    "am_email": am_email_c, "akcije": ac["akcije"]})
                                    save_promo_state_bulk(state_updates)
                                    save_promo_history(history_changes)
                                    if sent_log:
                                        append_alert_log(sent_log)
                                        log.info(f"[Scheduler-City] {city}: sent alerts for {len(sent_log)} restaurants.")
                    except Exception as e:
                        log.error(f"[Scheduler-City] {city}: error — {e}")
                    finally:
                        release_scan_lock()

        except Exception as e:
            import logging as _log
            _log.getLogger("scheduler").error(f"[Scheduler] loop error: {e}")

        time.sleep(30)

def start_global_background_threads():
    if not st.session_state.get("_scheduler_thread_started", False):
        t_sch = threading.Thread(target=_scheduler_loop, daemon=True)
        t_sch.daemon = True
        t_sch.start()
        st.session_state["_scheduler_thread_started"] = True

start_global_background_threads()

# ─────────────────────────── SESSION STATE ───────────────────────────────────

if "df_wolt" not in st.session_state:
    st.session_state.df_wolt = pd.DataFrame()
if "last_scan" not in st.session_state:
    st.session_state.last_scan = None
if "scan_stop_event" not in st.session_state:
    st.session_state.scan_stop_event = threading.Event()
if "scan_running" not in st.session_state:
    st.session_state.scan_running = False
if "scan_start_time" not in st.session_state:
    st.session_state.scan_start_time = None
if "scan_mode" not in st.session_state:
    st.session_state.scan_mode = "full"
if "scan_duration_last" not in st.session_state:
    st.session_state.scan_duration_last = None

if is_scan_locked() or Path("_scan_done.txt").exists():
    st.session_state.scan_running = True
else:
    st.session_state.scan_running = False

# ─────────────────────────── UI ──────────────────────────────────────────────

st.title("🏷️ Promo Monitor")
st.caption("Scans item-level discounts: checks each restaurant for at least one discounted product.")

tab_scan, tab_amm, tab_alert, tab_stats, tab_sched, tab_watchlist, tab_debug, tab_reset = st.tabs([
    "🔍 Scan & Results",
    "👥 AM Database",
    "📧 Send Alert",
    "📈 Statistics",
    "⏰ Scheduler",
    "⭐ Watchlist",
    "⚙️ Settings",
    "🗑️ Reset & Backup",
])

# ══════════════════════════════════════════════════════════════════════════════
# TAB 1: SCAN
# ══════════════════════════════════════════════════════════════════════════════
with tab_scan:
    scan_done_flag = Path("_scan_done.txt").exists()

    # DETEKCIJA ZAGLAVLJENOG ILI AKTIVNOG SKENIRANJA SA DUGMETOM ZA SPASAVANJE
    if st.session_state.scan_running and not scan_done_flag:
        try:
            with sqlite3.connect(DB_FILE, check_same_thread=False) as _lc:
                _lr = _lc.execute("SELECT locked_at FROM scan_lock WHERE id=1").fetchone()
            elapsed = time.time() - (_lr[0] if _lr else (st.session_state.scan_start_time or time.time()))
        except Exception:
            elapsed = time.time() - (st.session_state.scan_start_time or time.time())
        m2, s2 = divmod(int(elapsed), 60)
        
        st.markdown(f"### 🔄 Scan in progress — {m2:02d}:{s2:02d}")
        
        # DODATO CRVENO DUGME ZA PRILIKU KADA SE PROCES NASILNO UBIJE DA SE EKRAN ODGLAVI ODMAH
        if st.button("🚨 Force Stop / Odglavi ekran i katanac baze", type="secondary"):
            release_scan_lock()
            Path("_scan_done.txt").unlink(missing_ok=True)
            Path("_scan_status.txt").unlink(missing_ok=True)
            Path("_scan_city_progress.json").unlink(missing_ok=True)
            st.session_state.scan_running = False
            st.success("Sistem uspešno odglavljen!")
            st.rerun()
            
        city_prog = {}
        for _ in range(3):
            try:
                raw = Path("_scan_city_progress.json").read_text(encoding="utf-8")
                parsed = json.loads(raw)
                if isinstance(parsed, dict) and parsed:
                    city_prog = parsed
                    break
            except Exception:
                time.sleep(0.05)
        if city_prog:
            st.session_state["_last_city_prog"] = city_prog
        else:
            city_prog = st.session_state.get("_last_city_prog", {})
        if city_prog:
            PROG_PER_ROW = 6
            city_items = list(city_prog.items())
            for chunk_start in range(0, len(city_items), PROG_PER_ROW):
                chunk = city_items[chunk_start:chunk_start + PROG_PER_ROW]
                cols = st.columns(PROG_PER_ROW)
                for col_idx, (city_name, info) in enumerate(chunk):
                    with cols[col_idx]:
                        found  = info.get("found", 0)
                        cstatus = info.get("status", "...")
                        is_done = "✅" in cstatus
                        color = "#27ae60" if is_done else "#009de0"
                        st.markdown(f"""
                    <div style='background:#fff;border-radius:10px;padding:14px 16px;
                                box-shadow:0 2px 8px rgba(0,0,0,0.08);
                                border-top:4px solid {color};margin-bottom:8px'>
                      <div style='font-weight:800;font-size:1.1rem;color:{color}'>{city_name}</div>
                      <div style='font-size:1.8rem;font-weight:900;color:#222'>{found}</div>
                      <div style='font-size:0.75rem;color:#888'>restaurants found</div>
                      <div style='font-size:0.8rem;color:#555;margin-top:6px'>{cstatus}</div>
                    </div>
                    """, unsafe_allow_html=True)
        time.sleep(5)
        st.rerun()

    st.markdown("### 🔍 Scan parameters")
    selected_cities = st.multiselect("📍 Cities to scan:", options=CITIES, default=CITIES, key="selected_cities")

    prev_df_for_nopromo = st.session_state.df_wolt
    nopromo_available = not prev_df_for_nopromo.empty
    if nopromo_available:
        no_promo_count = len(prev_df_for_nopromo[
            prev_df_for_nopromo["grad"].isin(selected_cities) &
            (prev_df_for_nopromo["akcije"] == "-")
        ]) if selected_cities else 0
    else:
        no_promo_count = 0

    col_btn, col_btn2, col_btn3, col_stop, col_info = st.columns([1.2, 1.5, 1.5, 0.9, 2.4])
    with col_btn:
        run_scan = st.button("▶️ Full Scan", type="primary", use_container_width=True,
                             disabled=not selected_cities or st.session_state.scan_running)
    with col_btn2:
        run_nopromo = st.button(f"🔍 No Promo Scan ({no_promo_count})", use_container_width=True,
                                disabled=not selected_cities or st.session_state.scan_running or not nopromo_available or no_promo_count == 0)
    with col_btn3:
        run_city_rescan = st.button("🔄 City Rescan", use_container_width=True,
                                    disabled=not selected_cities or st.session_state.scan_running or not nopromo_available,
                                    help="Re-scans selected cities from scratch and merges results with the existing scan (other cities are preserved).")
    with col_stop:
        stop_scan = st.button("⏹️ Stop", use_container_width=True,
                              disabled=not st.session_state.scan_running, type="secondary")
    with col_info:
        if st.session_state.last_scan:
            st.info(f"⏱️ Last scan: **{st.session_state.last_scan}** | Total restaurants: **{len(st.session_state.df_wolt)}**")
        if not selected_cities:
            st.warning("Select at least one city.")

    prev_meta = scan_meta()
    if prev_meta and not st.session_state.scan_running:
        load_col, _ = st.columns([2, 4])
        with load_col:
            if st.button(f"📂 Load previous scan ({prev_meta})", use_container_width=True):
                prev_df = load_scan()
                if not prev_df.empty:
                    st.session_state.df_wolt = prev_df
                    st.session_state.last_scan = prev_meta
                    st.success(f"✅ Previous scan loaded – {len(prev_df)} restaurants.")
                    st.rerun()
                else:
                    st.error("Error loading file.")

    if stop_scan and st.session_state.scan_running:
        st.session_state.scan_stop_event.set()
        release_scan_lock()
        Path("_scan_done.txt").unlink(missing_ok=True)
        st.warning("⏹️ Stopping and releasing database lock...")
        st.rerun()

    if run_city_rescan and selected_cities and not st.session_state.scan_running and nopromo_available:
        if not acquire_scan_lock():
            st.error("⛔ A scan is already running. Try again later.")
        else:
            st.session_state.scan_stop_event = threading.Event()
            st.session_state.scan_running = True
            st.session_state.scan_mode = "city_rescan"
            st.session_state.scan_start_time = time.time()
            _cities_snap = list(selected_cities)
            _stop_ev_snap = st.session_state.scan_stop_event
            _prev_df_snap = st.session_state.df_wolt.copy()
            Path("_scan_done.txt").unlink(missing_ok=True)
            Path("_scan_result.json").unlink(missing_ok=True)
            with _city_progress_lock:
                _city_progress.clear()
                for _c in _cities_snap:
                    _city_progress[_c] = {"found": 0, "total": 0, "status": "⏳ Waiting..."}
            _write_status_file()

            def _run_city_rescan_bg():
                try:
                    class LivePH:
                        def info(self, msg, *a, **k): Path("_scan_status.txt").write_text(str(msg))
                        def warning(self, msg, *a, **k): Path("_scan_status.txt").write_text("⚠️ " + str(msg))
                        def success(self, msg, *a, **k): Path("_scan_status.txt").write_text("✅ " + str(msg))
                        def error(self, msg, *a, **k): Path("_scan_status.txt").write_text("❌ " + str(msg))
                        def empty(self, *a, **k): pass
                    new_city_df = scan_all_cities(_cities_snap, LivePH(), _stop_ev_snap)
                    if new_city_df is not None and not new_city_df.empty:
                        other_cities_df = _prev_df_snap[~_prev_df_snap["grad"].isin(_cities_snap)].copy()
                        merged_df = pd.concat([other_cities_df, new_city_df], ignore_index=True)
                        merged_df.to_json("_scan_result.json", orient="records", force_ascii=False)
                    elif not _stop_ev_snap.is_set():
                        _prev_df_snap.to_json("_scan_result.json", orient="records", force_ascii=False)
                finally:
                    release_scan_lock()
                    Path("_scan_done.txt").write_text("1")

            threading.Thread(target=_run_city_rescan_bg, daemon=True).start()
            st.rerun()

    if run_nopromo and selected_cities and not st.session_state.scan_running and nopromo_available:
        if not acquire_scan_lock():
            st.error("⛔ A scan is already running (another user or scheduled scan). Try again later.")
        else:
            st.session_state.scan_stop_event = threading.Event()
            st.session_state.scan_running = True
            st.session_state.scan_mode = "nopromo"
            st.session_state.scan_start_time = time.time()
            _cities_snap = list(selected_cities)
            _stop_ev_snap = st.session_state.scan_stop_event
            _prev_df_snap = st.session_state.df_wolt.copy()
            Path("_scan_done.txt").unlink(missing_ok=True)
            Path("_scan_result.json").unlink(missing_ok=True)
            with _city_progress_lock:
                _city_progress.clear()
                for _c in _cities_snap:
                    _cnt = len(_prev_df_snap[(_prev_df_snap["grad"] == _c) & (_prev_df_snap["akcije"] == "-")])
                    _city_progress[_c] = {"found": _cnt, "total": _cnt, "status": f"⏳ Waiting... ({_cnt} rest.)"}
            _write_status_file()

            def _run_nopromo_bg():
                try:
                    result = scan_nopromo_cities(_cities_snap, _prev_df_snap, _stop_ev_snap)
                    if result is not None and not result.empty:
                        result.to_json("_scan_result.json", orient="records", force_ascii=False)
                finally:
                    release_scan_lock()
                    Path("_scan_done.txt").write_text("1")

            threading.Thread(target=_run_nopromo_bg, daemon=True).start()
            st.rerun()

    if run_scan and selected_cities and not st.session_state.scan_running:
        if not acquire_scan_lock():
            st.error("⛔ A scan is already running (another user or scheduled scan). Try again later.")
        else:
            st.session_state.scan_stop_event = threading.Event()
            st.session_state.scan_running = True
            st.session_state.scan_mode = "full"
            st.session_state.scan_start_time = time.time()
            _cities_snap = list(selected_cities)
            _stop_ev_snap = st.session_state.scan_stop_event
            Path("_scan_done.txt").unlink(missing_ok=True)
            Path("_scan_result.json").unlink(missing_ok=True)
            with _city_progress_lock:
                _city_progress.clear()
                for _c in _cities_snap:
                    _city_progress[_c] = {"found": 0, "total": 0, "status": "⏳ Waiting..."}
            _write_status_file()

            def _run_scan_bg():
                try:
                    class LivePH:
                        def info(self, msg, *a, **k): Path("_scan_status.txt").write_text(str(msg))
                        def warning(self, msg, *a, **k): Path("_scan_status.txt").write_text("⚠️ " + str(msg))
                        def success(self, msg, *a, **k): Path("_scan_status.txt").write_text("✅ " + str(msg))
                        def error(self, msg, *a, **k): Path("_scan_status.txt").write_text("❌ " + str(msg))
                        def empty(self, *a, **k): pass
                    result = scan_all_cities(_cities_snap, LivePH(), _stop_ev_snap)
                    if result is not None and not result.empty:
                        result.to_json("_scan_result.json", orient="records", force_ascii=False)
                finally:
                    release_scan_lock()
                    Path("_scan_done.txt").write_text("1")

            threading.Thread(target=_run_scan_bg, daemon=True).start()
            st.rerun()

    if st.session_state.scan_running and scan_done_flag:
        Path("_scan_done.txt").unlink(missing_ok=True)
        st.session_state.scan_running = False
        scan_duration = time.time() - (st.session_state.scan_start_time or time.time()) if st.session_state.scan_start_time else 60
        _stop_ev = st.session_state.scan_stop_event
        try:
            df_result = pd.read_json("_scan_result.json", orient="records")
        except Exception:
            df_result = pd.DataFrame()
        if df_result is not None and not df_result.empty:
            st.session_state.df_wolt = df_result
            st.session_state.last_scan = local_now()
            st.session_state.scan_duration_last = scan_duration
            save_scan(df_result)
            novi_df = df_result[df_result["novo"] == "Da"].copy() if "novo" in df_result.columns else pd.DataFrame()
            if not novi_df.empty:
                sent_slugs    = load_sent_new_restaurants()
                amm_check     = load_amm()
                sales_cfg     = load_sales()
                novi_po_gradu = {}
                new_sent_slugs = set(sent_slugs)
                for _, row in novi_df.iterrows():
                    naziv = row.get("naziv", "")
                    grad  = row.get("grad", "")
                    slug  = row.get("slug", "")
                    norm  = normalize(naziv)
                    if slug in sent_slugs:
                        continue
                    has_am = False
                    if not amm_check.empty:
                        has_am = not amm_check[
                            (amm_check["restaurant_norm"] == norm) & (amm_check["city"] == grad)
                        ].empty
                    if not has_am:
                        if grad not in novi_po_gradu:
                            novi_po_gradu[grad] = []
                        novi_po_gradu[grad].append({"naziv": naziv, "slug": slug})
                        new_sent_slugs.add(slug)
                notified = 0
                for grad_key, restorani in novi_po_gradu.items():
                    for email in sales_cfg.get(grad_key, []):
                        if send_sales_bulk_notification(email, grad_key, restorani):
                            notified += 1
                save_sent_new_restaurants(new_sent_slugs)
                if notified:
                    total_novi = sum(len(v) for v in novi_po_gradu.values())
                    st.info(f"📬 Sent **{notified}** bulk notifications to sales agents ({total_novi} new restaurants).")
            m, s = divmod(int(scan_duration), 60)
            scan_mode_done = st.session_state.get("scan_mode", "full")
            if scan_mode_done == "nopromo":
                newly_found = len(df_result[df_result["akcije"] != "-"])
                st.success(f"✅ No Promo Scan finished! Of previously skipped restaurants, **{newly_found}** now have promotions.")
            elif scan_mode_done == "city_rescan":
                st.success(f"✅ City Rescan finished! Total: **{len(df_result)}** restaurants, **{len(df_result[df_result['akcije'] != '-'])}** with promotions.")
            else:
                st.success(f"✅ Full Scan finished in **{m:02d}:{s:02d}**! Found **{len(df_result)}** restaurants, **{len(df_result[df_result['akcije'] != '-'])}** with promotions.")
            st.rerun()
        else:
            if _stop_ev.is_set():
                st.warning("⏹️ Scan was stopped.")
            else:
                st.error("❌ Scan returned no data.")

    df = st.session_state.df_wolt
    if not df.empty:
        if st.session_state.scan_duration_last:
            m_t, s_t = divmod(int(st.session_state.scan_duration_last), 60)
            st.markdown(f"<div style='background:#e8f8f0;border-left:4px solid #27ae60;padding:8px 16px;border-radius:6px;margin-bottom:12px;font-size:0.95rem;color:#155724'>⏱️ Last scan duration: <strong>{m_t:02d}:{s_t:02d}</strong></div>", unsafe_allow_html=True)
        st.markdown("---")

        k1, k2, k3, k4, k5 = st.columns(5)
        total        = len(df)
        sa_akcijama  = len(df[df["akcije"] != "-"])
        otvoreni     = len(df[df["status"] == "Otvoren"])
        novi         = len(df[df["novo"] == "Da"])
        sa_wolt_plus = len(df[df["akcije"].apply(lambda c: bool(re.search(r'\[Wolt\+\]|Wolt\+|W\+', c, re.IGNORECASE)) if pd.notna(c) else False)])

        for col, val, lbl in [
            (k1, total, "Total restaurants"), (k2, sa_akcijama, "Has promotion"),
            (k3, sa_wolt_plus, "💙 Wolt+ promotions"),
            (k4, otvoreni, "Currently open"), (k5, novi, "New restaurants"),
        ]:
            with col:
                st.markdown(f"<div class='kpi'><div class='kpi-val'>{val}</div><div class='kpi-lbl'>{lbl}</div></div>", unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)
        grad_summary = df.groupby("grad").agg(
            Restaurants=("naziv", "count"),
            With_promos=("akcije", lambda x: (x != "-").sum()),
            Open=("status", lambda x: (x == "Otvoren").sum()),
        ).reset_index()
        CARDS_PER_ROW = 6
        rows_data = [grad_summary.iloc[i:i+CARDS_PER_ROW] for i in range(0, len(grad_summary), CARDS_PER_ROW)]
        for chunk in rows_data:
            gs_cols = st.columns(CARDS_PER_ROW)
            for col_idx, (_, row) in enumerate(chunk.iterrows()):
                with gs_cols[col_idx]:
                    pct = int(row["With_promos"] / row["Restaurants"] * 100) if row["Restaurants"] > 0 else 0
                    st.markdown(f"""
                <div style='background:#fff;border-radius:10px;padding:12px 16px;box-shadow:0 2px 8px rgba(0,0,0,0.07);border-top:3px solid #009de0;text-align:center'>
                  <div style='font-weight:800;color:#009de0;font-size:1rem'>{row["grad"]}</div>
                  <div style='font-size:1.6rem;font-weight:900'>{int(row["Restaurants"])}</div>
                  <div style='font-size:0.75rem;color:#888'>restaurants</div>
                  <div style='margin-top:4px;font-size:0.85rem;color:#27ae60'>{int(row["With_promos"])} promos ({pct}%)</div>
                  <div style='font-size:0.75rem;color:#555'>{int(row["Open"])} open</div>
                </div>""", unsafe_allow_html=True)
            st.markdown("<br>", unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)

        ff1, ff2, ff3, ff4, ff5, ff6 = st.columns([2, 1, 1, 1, 1, 2])
        with ff1: grad_filter = st.multiselect("📍 City:", CITIES, default=CITIES, key="scan_grad")
        with ff2: samo_akcije = st.checkbox("📌 With promos", value=False, key="scan_akcije")
        with ff3: samo_wolt_plus = st.checkbox("💙 Wolt+", value=False, key="scan_wolt_plus")
        with ff4: samo_pct_popust = st.checkbox("🔢 % discount", value=False, key="scan_pct_popust")
        with ff5: samo_otvoreni = st.checkbox("🟢 Open", value=False, key="scan_otvoreni")
        with ff6: search = st.text_input("🔎 Search name:", key="scan_search", placeholder="restaurant name...")

        ff7, ff8 = st.columns([1, 3])
        with ff7: samo_novi = st.checkbox("🆕 New only", value=False, key="scan_novi")
        with ff8:
            sve_akcije_tekst = sorted(set(
                line.lstrip("• ").strip()
                for akcije_cell in df["akcije"] if akcije_cell != "-"
                for line in akcije_cell.split("\n") if line.strip() and line.strip() != "-"
            ))
            akcija_filter = st.multiselect("🎯 Filter by promo type:", options=sve_akcije_tekst, default=[], key="scan_akcija_filter", placeholder="All promotions...")

        fdf = df[df["grad"].isin(grad_filter)]
        if samo_akcije: fdf = fdf[fdf["akcije"] != "-"]
        if samo_novi: fdf = fdf[fdf["novo"] == "Da"]
        if samo_otvoreni: fdf = fdf[fdf["status"] == "Otvoren"]
        if search.strip(): fdf = fdf[fdf["naziv"].str.contains(search.strip(), case=False, na=False)]
        if akcija_filter:
            fdf = fdf[fdf["akcije"].apply(lambda cell: any(a in cell for a in akcija_filter) if cell != "-" else False)]
        if samo_wolt_plus:
            fdf = fdf[fdf["akcije"].apply(lambda cell: bool(re.search(r'\[Wolt\+\]|Wolt\+|W\+', cell, re.IGNORECASE)) if cell != "-" else False)]
        if samo_pct_popust:
            fdf = fdf[fdf["akcije"].str.contains(r'\d+\s*%', na=False, regex=True)]

        display_cols = ["grad", "naziv", "status", "ocena", "dostava", "novo", "akcije", "link"]
        display_cols = [c for c in display_cols if c in fdf.columns]
        st.dataframe(fdf[display_cols].reset_index(drop=True), use_container_width=True, hide_index=True, height=480,
            column_config={
                "grad": st.column_config.TextColumn("City"), "naziv": st.column_config.TextColumn("Restaurant"),
                "status": st.column_config.TextColumn("Status"), "ocena": st.column_config.TextColumn("Rating"),
                "dostava": st.column_config.TextColumn("Delivery"), "novo": st.column_config.TextColumn("New"),
                "akcije": st.column_config.TextColumn("Promotions", width="large"),
                "link": st.column_config.LinkColumn("Link", display_text="Open ↗"),
            })
        csv = fdf[display_cols].to_csv(index=False).encode("utf-8")
        st.download_button("📥 Download CSV", csv, "scan.csv", "text/csv")

# ══════════════════════════════════════════════════════════════════════════════
# TAB 2: AMM BAZA
# ══════════════════════════════════════════════════════════════════════════════
with tab_amm:
    st.markdown("### 👥 Account Manager Database")
    st.caption("Define which AM is responsible for which restaurant. Stored in local SQLite database.")

    amm_df  = load_amm()
    df_wolt = st.session_state.df_wolt

    st.markdown("---")
    st.markdown("#### 📬 Sales agents by city")
    sales_data = load_sales()

    for city in CITIES:
        emails_current = sales_data.get(city, [])
        col_city, col_email, col_save = st.columns([1, 3, 1])
        with col_city: st.markdown(f"**{city}**")
        with col_email:
            new_emails_str = st.text_input(f"Email(ovi) za {city}:", value=", ".join(emails_current),
                                           placeholder="sales@firma.com", key=f"sales_email_{city}", label_visibility="collapsed")
        with col_save:
            if st.button("💾", key=f"sales_save_{city}"):
                parsed = [e.strip() for e in new_emails_str.split(",") if e.strip()]
                sales_data[city] = parsed
                save_sales(sales_data)
                st.success(f"✅ {city}: {len(parsed)} email(s) saved.")
                st.rerun()

    st.markdown("---")
    st.markdown("#### ⚡ Bulk assign")

    if df_wolt.empty:
        st.info("Run a scan first to see restaurants.")
    else:
        bulk_am_opts = sorted(amm_df["am_name"].dropna().unique().tolist()) if not amm_df.empty else []
        if not bulk_am_opts:
            st.warning("No AMs in the database. Add an AM below and come back.")
        else:
            b_col1, b_col2 = st.columns([1, 3])
            with b_col1: bulk_selected_am = st.selectbox("Select AM:", bulk_am_opts, key="bulk_am_sel")
            with b_col2: bulk_grad = st.multiselect("Filter by city:", CITIES, default=CITIES, key="bulk_grad_filt")

            if bulk_selected_am:
                am_row = amm_df[amm_df["am_name"] == bulk_selected_am].iloc[0]
                am_email_bulk = am_row["am_email"]

                # Svi restorani iz skena za izabrane gradove
                bulk_df = df_wolt[df_wolt["grad"].isin(bulk_grad)][["naziv", "grad"]].drop_duplicates().copy()

                # Skup (norm, city) parova koji već imaju STVARNOG AM-a (am_name nije prazan)
                amm_sa_am = amm_df[amm_df["am_name"].notna() & (amm_df["am_name"].str.strip() != "") & (amm_df["am_name"] != "None")]
                already_assigned = set(
                    zip(amm_sa_am["restaurant_norm"].apply(lambda x: str(x).strip()),
                        amm_sa_am["city"].apply(lambda x: str(x).strip()))
                )

                # Prikaži samo one koji NEMAJU nijednog AM-a (po norm + grad)
                bulk_df["_norm"] = bulk_df["naziv"].apply(normalize)
                bulk_df["_grad"] = bulk_df["grad"].apply(lambda x: str(x).strip())
                bulk_df = bulk_df[
                    ~bulk_df.apply(lambda r: (r["_norm"], r["_grad"]) in already_assigned, axis=1)
                ].drop(columns=["_norm", "_grad"])
                bulk_df["✅ Assign"] = False

                if bulk_df.empty:
                    st.info("✅ Svi restorani u izabranim gradovima već imaju dodeljenog AM-a.")
                else:
                    st.caption(f"Prikazano **{len(bulk_df)}** restorana bez AM-a.")
                    edited_bulk = st.data_editor(bulk_df.reset_index(drop=True), use_container_width=True,
                        hide_index=True, height=400,
                        column_config={
                            "✅ Assign": st.column_config.CheckboxColumn("Assign to this AM", default=False),
                            "naziv": st.column_config.TextColumn("Restaurant", disabled=True),
                            "grad": st.column_config.TextColumn("City", disabled=True),
                        }, key="bulk_editor")

                if st.button("💾 Save bulk assignment", key="bulk_save"):
                        selected_rows = edited_bulk[edited_bulk["✅ Assign"] == True]
                        new_rows = []
                        for _, row in selected_rows.iterrows():
                            norm = normalize(row["naziv"])
                            city_v = row["grad"]
                            mask = (amm_df["restaurant_norm"] == norm) & (amm_df["city"] == city_v)
                            if mask.any():
                                amm_df.loc[mask, ["am_name", "am_email"]] = [bulk_selected_am, am_email_bulk]
                            else:
                                new_rows.append({"restaurant_norm": norm, "restaurant_display": row["naziv"],
                                                "city": city_v, "am_name": bulk_selected_am, "am_email": am_email_bulk})
                        if new_rows:
                            amm_df = pd.concat([amm_df, pd.DataFrame(new_rows)], ignore_index=True)
                        save_amm(amm_df)
                        st.success(f"✅ Assigned {len(selected_rows)} restaurants → {bulk_selected_am}")
                        st.rerun()

    st.markdown("---")
    st.markdown("#### ➕ Add / update individually")

    rest_options = sorted(df_wolt["naziv"].dropna().unique().tolist()) if not df_wolt.empty else []
    a1, a2 = st.columns([2, 1])
    with a1: sel_rest = st.selectbox("Restaurant:", ["-- Select --"] + rest_options, key="amm_sel")
    with a2: man_rest = st.text_input("Or type manually:", placeholder="e.g. KFC", key="amm_man")

    final_rest = man_rest.strip() if man_rest.strip() else (sel_rest if sel_rest != "-- Select --" else "")
    b1, b2, b3, b4 = st.columns(4)
    with b1: amm_city  = st.selectbox("City:", ["-- All --"] + CITIES, key="amm_city_sel")
    with b2: amm_name  = st.text_input("AM Name:", key="amm_name")
    with b3: amm_email = st.text_input("AM Email:", key="amm_email")
    with b4:
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("💾 Save", use_container_width=True, key="amm_save"):
            if not final_rest:
                st.error("Select or type a restaurant name.")
            elif not amm_name or not amm_email:
                st.error("Enter AM name and email.")
            else:
                norm = normalize(final_rest)
                city_val = "" if amm_city == "-- All --" else amm_city
                mask = (amm_df["restaurant_norm"] == norm) & (amm_df["city"] == city_val)
                if mask.any():
                    amm_df.loc[mask, ["restaurant_display", "am_name", "am_email"]] = [final_rest, amm_name, amm_email]
                else:
                    amm_df = pd.concat([amm_df, pd.DataFrame([{"restaurant_norm": norm, "restaurant_display": final_rest,
                                        "city": city_val, "am_name": amm_name, "am_email": amm_email}])], ignore_index=True)
                save_amm(amm_df)
                st.success(f"✅ Saved to Local Database: **{final_rest}** → {amm_name}")
                st.rerun()

    st.markdown("---")
    st.markdown("#### 📋 Current database")
    if amm_df.empty:
        st.info("Database is empty.")
    else:
        am_opts = ["All"] + sorted(amm_df["am_name"].dropna().unique().tolist())
        am_filt = st.selectbox("Filter by AM:", am_opts, key="amm_view_filt")
        view = amm_df if am_filt == "All" else amm_df[amm_df["am_name"] == am_filt]
        edited = st.data_editor(view.reset_index(drop=True), use_container_width=True, num_rows="dynamic",
            hide_index=True, column_config={
                "restaurant_norm": st.column_config.TextColumn("Norm name", disabled=True),
                "restaurant_display": st.column_config.TextColumn("Restaurant"),
                "city": st.column_config.TextColumn("City"),
                "am_name": st.column_config.TextColumn("AM Name"),
                "am_email": st.column_config.TextColumn("AM Email"),
            }, key="amm_editor")
        if st.button("💾 Save changes", key="amm_save_tbl"):
            if am_filt == "All":
                save_amm(edited)
            else:
                rest_df = amm_df[amm_df["am_name"] != am_filt]
                save_amm(pd.concat([rest_df, edited], ignore_index=True))
            st.success("✅ Database updated in SQLite!")
            st.rerun()

    st.markdown("---")
    st.markdown("#### 📥 Bulk import CSV")
    uploaded = st.file_uploader("CSV file:", type="csv", key="amm_upload")
    if uploaded:
        try:
            new_df = pd.read_csv(uploaded)
            new_df["restaurant_norm"] = new_df["restaurant_display"].apply(normalize)
            merged_amm = pd.concat([amm_df, new_df], ignore_index=True).drop_duplicates(subset=["restaurant_norm", "city"], keep="last")
            save_amm(merged_amm)
            st.success(f"✅ Imported {len(new_df)} rows.")
            st.rerun()
        except Exception as e:
            st.error(f"Error: {e}")

# ══════════════════════════════════════════════════════════════════════════════
# TAB 3: POŠALJI ALERT
# ══════════════════════════════════════════════════════════════════════════════
with tab_alert:
    st.markdown("### 📧 Send Alert to AMs")
    df_wolt = st.session_state.df_wolt
    amm_df  = load_amm()

    if df_wolt.empty:
        st.warning("⚠️ No scan data available.")
    elif amm_df.empty:
        st.warning("⚠️ AM database is empty.")
    else:
        df_wolt["naziv_norm"] = df_wolt["naziv"].apply(normalize)
        merged = df_wolt.merge(
            amm_df[["restaurant_norm", "restaurant_display", "city", "am_name", "am_email"]],
            left_on="naziv_norm", right_on="restaurant_norm", how="inner"
        )
        merged["_alert"] = merged.apply(lambda row: filter_akcije_for_email(row["akcije"]) != "-", axis=1)
        sa_akcijama = merged[merged["_alert"]].copy()

        if sa_akcijama.empty:
            st.info("✅ None of your partners currently have relevant promotions.")
        else:
            sa_akcijama["akcije_email"] = sa_akcijama["akcije"].apply(filter_akcije_for_email)
            af1, af2 = st.columns(2)
            with af1: grad_filt_a = st.multiselect("City:", CITIES, default=CITIES, key="alert_grad")
            with af2: am_filt_a = st.multiselect("AM:", sorted(sa_akcijama["am_name"].dropna().unique().tolist()),
                                                  default=sorted(sa_akcijama["am_name"].dropna().unique().tolist()), key="alert_am")
            preview = sa_akcijama[(sa_akcijama["grad"].isin(grad_filt_a)) & (sa_akcijama["am_name"].isin(am_filt_a))]
            st.caption(f"Partners to alert: **{len(preview)}** | AMs: **{preview['am_name'].nunique()}**")
            preview_cols = ["grad", "naziv", "am_name", "am_email", "akcije_email", "link"]
            preview_cols = [c for c in preview_cols if c in preview.columns]
            st.dataframe(preview[preview_cols].reset_index(drop=True), use_container_width=True, hide_index=True, height=350,
                column_config={
                    "grad": st.column_config.TextColumn("City"), "naziv": st.column_config.TextColumn("Restaurant"),
                    "am_name": st.column_config.TextColumn("AM"), "am_email": st.column_config.TextColumn("Email"),
                    "akcije_email": st.column_config.TextColumn("Promotions", width="large"),
                    "link": st.column_config.LinkColumn("Link", display_text="Open ↗"),
                })
            st.markdown("---")
            if st.button("🚀 Send alerts", type="primary"):
                promo_state = load_promo_state()
                cooldown    = load_alert_cooldown()
                state_updates = {}
                history_changes = []
                am_groups = preview.groupby(["am_name", "am_email"])
                sent_log = []
                success_count = 0
                skipped_count = 0
                cooldown_count = 0
                for (am_name, am_email_addr), grp in am_groups:
                    alerts = []
                    for _, row in grp.iterrows():
                        rest_norm = normalize(row["naziv"])
                        city      = row["grad"]
                        akcije_filtered = filter_akcije_for_email(row["akcije"])
                        prev_state = promo_state.get((rest_norm, city))
                        state_updates[(rest_norm, city)] = akcije_filtered
                        # Zabelezi promenu u istoriju
                        if prev_state is not None and prev_state != akcije_filtered:
                            history_changes.append({
                                "restaurant_norm":    rest_norm,
                                "restaurant_display": row["naziv"],
                                "city":               city,
                                "stare_akcije":       prev_state,
                                "nove_akcije":        akcije_filtered,
                            })
                        if not should_send_am_alert(rest_norm, city, akcije_filtered, promo_state):
                            skipped_count += 1
                            continue
                        if is_in_cooldown(am_email_addr, rest_norm, cooldown):
                            cooldown_count += 1
                            continue
                        alerts.append({
                            "naziv": row["naziv"], "grad": city,
                            "akcije": row["akcije"], "link": row.get("link", ""),
                            "norm": rest_norm,
                        })
                    if not alerts:
                        continue
                    all_partners = grp[["naziv", "grad", "akcije", "link"]].to_dict("records") if all(c in grp.columns for c in ["naziv","grad","akcije","link"]) else []
                    ok = send_alert_email(am_email_addr, am_name, alerts, all_partners=all_partners, tip="AM-Manualno")
                    if ok:
                        success_count += 1
                        for a in alerts:
                            update_cooldown(am_email_addr, a["norm"], cooldown)
                        st.success(f"✅ Email sent: **{am_name}** – {len(alerts)} partners")
                        for a in alerts:
                            sent_log.append({"timestamp": local_now(), "city": a["grad"],
                                            "restaurant_display": a["naziv"], "am_name": am_name,
                                            "am_email": am_email_addr, "akcije": a["akcije"]})
                    else:
                        st.error(f"❌ Error: {am_name}")
                save_promo_state_bulk(state_updates)
                save_promo_history(history_changes)
                save_alert_cooldown(cooldown)
                if sent_log:
                    append_alert_log(sent_log)
                if skipped_count:
                    st.info(f"ℹ️ {skipped_count} partnera preskočeno (iste akcije kao pri poslednjem slanju).")
                if cooldown_count:
                    st.info(f"⏳ {cooldown_count} partnera preskočeno (cooldown {COOLDOWN_DAYS} dana).")
                st.markdown(f"**Done:** {success_count}/{am_groups.ngroups} AMs contacted.")

        st.markdown("---")
        st.markdown("#### 📋 Nedeljni Digest")
        st.caption("Šalje svakom AM-u pregled SVIH njegovih partnera — bez obzira na promene ili cooldown. Korisno za redovni nedeljni pregled.")
        if st.button("📋 Pošalji Nedeljni Digest svim AM-ovima", key="digest_btn"):
            all_merged = df_wolt.copy()
            all_merged["naziv_norm"] = all_merged["naziv"].apply(normalize)
            full_merged = all_merged.merge(
                amm_df[["restaurant_norm", "city", "am_name", "am_email"]],
                left_on="naziv_norm", right_on="restaurant_norm", how="inner"
            )
            if full_merged.empty:
                st.warning("Nema podataka za slanje.")
            else:
                digest_success = 0
                for (am_n, am_e), grp_d in full_merged.groupby(["am_name", "am_email"]):
                    partners_d = grp_d[["naziv", "grad", "akcije", "link"]].to_dict("records")
                    ok_d = send_digest_email(am_e, am_n, partners_d)
                    if ok_d:
                        digest_success += 1
                        st.success(f"✅ Digest poslat: **{am_n}** ({len(partners_d)} partnera)")
                    else:
                        st.error(f"❌ Greška: {am_n}")
                st.markdown(f"**Digest završen:** {digest_success} AM-ova kontaktirano.")

# ══════════════════════════════════════════════════════════════════════════════
# TAB 4: STATISTIKA — PUNI DASHBOARD
# ══════════════════════════════════════════════════════════════════════════════
with tab_stats:
    st.markdown("### 📈 Analytics Dashboard")

    log_df = load_alert_log()

    # ── Globalni KPI ──────────────────────────────────────────────────────────
    total_sent  = len(log_df)
    total_rest  = log_df["restaurant_display"].nunique() if not log_df.empty else 0
    total_am    = log_df["am_name"].nunique() if not log_df.empty else 0
    total_cities = log_df["city"].nunique() if not log_df.empty else 0

    k1, k2, k3, k4 = st.columns(4)
    for col, val, lbl, color in [
        (k1, total_sent,   "Alertova poslato",   "#009de0"),
        (k2, total_rest,   "Restorana praćeno",  "#27ae60"),
        (k3, total_am,     "Aktivnih AM-ova",    "#e67e22"),
        (k4, total_cities, "Gradova",            "#8e44ad"),
    ]:
        with col:
            st.markdown(
                f"<div class='kpi' style='border-top:4px solid {color}'>"
                f"<div class='kpi-val' style='color:{color}'>{val}</div>"
                f"<div class='kpi-lbl'>{lbl}</div></div>",
                unsafe_allow_html=True
            )

    st.markdown("<br>", unsafe_allow_html=True)

    # ── Filteri ───────────────────────────────────────────────────────────────
    cf1, cf2 = st.columns(2)
    with cf1:
        date_from_s = st.date_input("Od:", datetime.date.today() - datetime.timedelta(days=30), key="stats_from")
    with cf2:
        date_to_s = st.date_input("Do:", datetime.date.today(), key="stats_to")

    st.markdown("---")

    # ── SEKCIJA 1: AM Alert statistika ───────────────────────────────────────
    st.markdown("#### 👥 Alert statistika po AM-u")

    if log_df.empty:
        st.info("Još nema poslatih alertova.")
    else:
        ldf = log_df.copy()
        ldf["timestamp"] = pd.to_datetime(ldf["timestamp"], errors="coerce")
        ldf = ldf[ldf["timestamp"].dt.date >= date_from_s]
        ldf = ldf[ldf["timestamp"].dt.date <= date_to_s]

        if ldf.empty:
            st.warning("Nema alertova za izabrani period.")
        else:
            am_stats = (ldf.groupby(["am_name", "am_email"])
                        .agg(Alertova=("timestamp","count"),
                             Restorana=("restaurant_display","nunique"),
                             Gradova=("city","nunique"),
                             Poslednji=("timestamp","max"))
                        .reset_index()
                        .rename(columns={"am_name":"AM","am_email":"Email"})
                        .sort_values("Alertova", ascending=False))
            am_stats["Poslednji"] = am_stats["Poslednji"].dt.strftime("%d.%m.%Y %H:%M")
            st.dataframe(am_stats, use_container_width=True, hide_index=True)

            st.markdown("<br>", unsafe_allow_html=True)
            am_chart = am_stats.set_index("AM")["Alertova"]
            st.markdown("**Broj alertova po AM-u:**")
            st.bar_chart(am_chart)

    st.markdown("---")

    # ── SEKCIJA 2: Istorija promo promena ────────────────────────────────────
    st.markdown("#### 🔄 Istorija promo promena")
    st.caption("Svaki put kada se akcija promeni (nova, izmenjena, uklonjena) — beleži se ovde.")

    hist_days = st.slider("Prikaži istoriju za poslednjih N dana:", 7, 90, 30, key="hist_days_slider")
    hist_df = load_promo_history(days=hist_days)

    if hist_df.empty:
        st.info("Još nema zabeleženih promena akcija. Promene se automatski beleže pri svakom slanju AM alertova.")
    else:
        hf1, hf2 = st.columns(2)
        with hf1:
            hist_search = st.text_input("🔎 Pretraži restoran:", key="hist_search")
        with hf2:
            hist_city = st.selectbox("Grad:", ["Svi"] + CITIES, key="hist_city_f")

        fhist = hist_df.copy()
        if hist_search:
            fhist = fhist[fhist["restaurant_display"].str.contains(hist_search, case=False, na=False)]
        if hist_city != "Svi":
            fhist = fhist[fhist["city"] == hist_city]

        st.caption(f"Ukupno promena u periodu: **{len(fhist)}**")

        display_hist = fhist[["timestamp", "restaurant_display", "city", "stare_akcije", "nove_akcije"]].copy()
        display_hist.columns = ["Vreme", "Restoran", "Grad", "Stare akcije", "Nove akcije"]
        st.dataframe(display_hist.reset_index(drop=True), use_container_width=True, hide_index=True, height=400,
            column_config={
                "Stare akcije": st.column_config.TextColumn(width="large"),
                "Nove akcije": st.column_config.TextColumn(width="large"),
            })

        st.download_button(
            "📥 Export istorije (CSV)",
            display_hist.to_csv(index=False).encode("utf-8"),
            f"promo_istorija_{datetime.date.today()}.csv",
            "text/csv",
            key="hist_export"
        )

        # Grafikon — broj promena po danu (line chart)
        fhist_copy = fhist.copy()
        fhist_copy["datum"] = pd.to_datetime(fhist_copy["timestamp"]).dt.date
        daily_changes = fhist_copy.groupby("datum").size().reset_index(name="Promena")
        if not daily_changes.empty:
            try:
                import plotly.graph_objects as go
                fig_line = go.Figure()
                fig_line.add_trace(go.Scatter(
                    x=daily_changes["datum"], y=daily_changes["Promena"],
                    mode="lines+markers", fill="tozeroy",
                    line=dict(color="#009de0", width=2),
                    marker=dict(size=6, color="#009de0"),
                    name="Promene"
                ))
                fig_line.update_layout(
                    title="Broj promena akcija po danu",
                    xaxis_title="Datum", yaxis_title="Broj promena",
                    plot_bgcolor="white", paper_bgcolor="white",
                    height=300, margin=dict(l=10, r=10, t=40, b=10),
                )
                fig_line.update_xaxes(showgrid=True, gridcolor="#f0f0f0")
                fig_line.update_yaxes(showgrid=True, gridcolor="#f0f0f0")
                st.plotly_chart(fig_line, use_container_width=True)
            except ImportError:
                st.bar_chart(daily_changes.set_index("datum")["Promena"])

    st.markdown("---")

    # ── SEKCIJA 3: Istorija akcija po restoranu (alert log) ──────────────────
    st.markdown("#### 🍽️ Istorija alertovanih akcija po restoranu")

    if log_df.empty:
        st.info("Nema podataka.")
    else:
        ldf2 = log_df.copy()
        ldf2["timestamp"] = pd.to_datetime(ldf2["timestamp"], errors="coerce")
        ldf2 = ldf2[ldf2["timestamp"].dt.date >= date_from_s]
        ldf2 = ldf2[ldf2["timestamp"].dt.date <= date_to_s]

        if not ldf2.empty:
            rest_stats = (ldf2.groupby(["restaurant_display","city"])
                          .agg(Puta_u_promo=("timestamp","count"),
                               Prvih_detektovano=("timestamp","min"),
                               Poslednji_put=("timestamp","max"))
                          .reset_index()
                          .rename(columns={"restaurant_display":"Restoran","city":"Grad"})
                          .sort_values("Puta_u_promo", ascending=False))
            rest_stats["Prvih_detektovano"] = rest_stats["Prvih_detektovano"].dt.strftime("%d.%m.%Y")
            rest_stats["Poslednji_put"] = rest_stats["Poslednji_put"].dt.strftime("%d.%m.%Y")

            rc1, rc2 = st.columns([3,1])
            with rc1:
                rest_search = st.text_input("🔎 Pretraži restoran:", key="rest_search")
            with rc2:
                city_filter_r = st.selectbox("Grad:", ["Svi"] + CITIES, key="rest_city_f")

            filtered_rest = rest_stats.copy()
            if rest_search:
                filtered_rest = filtered_rest[filtered_rest["Restoran"].str.contains(rest_search, case=False, na=False)]
            if city_filter_r != "Svi":
                filtered_rest = filtered_rest[filtered_rest["Grad"] == city_filter_r]

            st.dataframe(filtered_rest, use_container_width=True, hide_index=True, height=400)

            try:
                import plotly.graph_objects as go
                top10 = rest_stats.head(10)
                fig_top = go.Figure(go.Bar(
                    x=top10["Puta_u_promo"], y=top10["Restoran"],
                    orientation="h", marker_color="#009de0",
                    text=top10["Puta_u_promo"], textposition="outside"
                ))
                fig_top.update_layout(
                    title="Top 10 restorana po broju promo detekcija",
                    xaxis_title="Broj detekcija", yaxis=dict(autorange="reversed"),
                    plot_bgcolor="white", paper_bgcolor="white",
                    height=350, margin=dict(l=10, r=40, t=40, b=10),
                )
                fig_top.update_xaxes(showgrid=True, gridcolor="#f0f0f0")
                st.plotly_chart(fig_top, use_container_width=True)
            except ImportError:
                st.bar_chart(rest_stats.head(10).set_index("Restoran")["Puta_u_promo"])

    st.markdown("---")

    # ── SEKCIJA 4: Trend po gradu + Wolt+ vs Regular pie ─────────────────────
    st.markdown("#### 🏙️ Aktivnost po gradu")

    if not log_df.empty:
        try:
            import plotly.graph_objects as go
            ldf3 = log_df.copy()
            ldf3["timestamp"] = pd.to_datetime(ldf3["timestamp"], errors="coerce")
            ldf3 = ldf3[ldf3["timestamp"].dt.date >= date_from_s]
            ldf3 = ldf3[ldf3["timestamp"].dt.date <= date_to_s]
            if not ldf3.empty:
                city_stats = (ldf3.groupby("city")
                              .agg(Alertova=("timestamp","count"),
                                   Restorana=("restaurant_display","nunique"))
                              .reset_index()
                              .rename(columns={"city":"Grad"})
                              .sort_values("Alertova", ascending=False))
                cc1, cc2 = st.columns(2)
                with cc1:
                    fig_city = go.Figure(go.Bar(
                        x=city_stats["Grad"], y=city_stats["Alertova"],
                        marker_color="#009de0",
                        text=city_stats["Alertova"], textposition="outside"
                    ))
                    fig_city.update_layout(
                        title="Alertova po gradu", xaxis_tickangle=-45,
                        plot_bgcolor="white", paper_bgcolor="white",
                        height=350, margin=dict(l=10, r=10, t=40, b=80),
                    )
                    fig_city.update_yaxes(showgrid=True, gridcolor="#f0f0f0")
                    st.plotly_chart(fig_city, use_container_width=True)
                with cc2:
                    # Pie chart: Wolt+ vs Regular promo iz tekućeg skena
                    df_pie = st.session_state.df_wolt
                    if not df_pie.empty and "akcije" in df_pie.columns:
                        has_wp = df_pie["akcije"].apply(
                            lambda c: bool(re.search(r'\[Wolt\+\]|Wolt\+|W\+', c, re.IGNORECASE)) if pd.notna(c) and c != "-" else False
                        )
                        has_reg = df_pie["akcije"].apply(
                            lambda c: (not re.search(r'\[Wolt\+\]|Wolt\+|W\+', c, re.IGNORECASE)) if (pd.notna(c) and c != "-") else False
                        )
                        n_wp  = has_wp.sum()
                        n_reg = has_reg.sum()
                        n_no  = (df_pie["akcije"] == "-").sum()
                        if n_wp + n_reg + n_no > 0:
                            fig_pie = go.Figure(go.Pie(
                                labels=["Wolt+", "Regular promo", "Bez promo"],
                                values=[n_wp, n_reg, n_no],
                                marker_colors=["#009de0", "#27ae60", "#e0e0e0"],
                                hole=0.4,
                                textinfo="label+percent",
                            ))
                            fig_pie.update_layout(
                                title="Wolt+ vs Regular vs Bez promo",
                                plot_bgcolor="white", paper_bgcolor="white",
                                height=350, margin=dict(l=10, r=10, t=40, b=10),
                                showlegend=False,
                            )
                            st.plotly_chart(fig_pie, use_container_width=True)
                    else:
                        st.info("Pokreni sken za prikaz pie charta.")
        except Exception:
            pass

    st.markdown("---")

    # ── SEKCIJA 5: Export — PDF i Excel jednim klikom ─────────────────────────
    st.markdown("#### 📄 Export Izveštaja")
    st.caption("PDF za menadžment ili Excel po gradovima sa bojama za promo status.")

    exp_col1, exp_col2 = st.columns(2)
    with exp_col1:
        pdf_city_filter = st.multiselect("Gradovi za export:", CITIES, default=CITIES, key="pdf_city_f")

    # ── Excel Export ──────────────────────────────────────────────────────────
    with exp_col2:
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("📊 Generiši Excel izveštaj", key="excel_gen_btn"):
            df_za_excel = st.session_state.df_wolt
            if df_za_excel.empty:
                st.warning("⚠️ Nema podataka skena. Najpre uradi scan.")
            else:
                try:
                    import io
                    import openpyxl
                    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
                    from openpyxl.utils import get_column_letter

                    fdf_xl = df_za_excel[df_za_excel["grad"].isin(pdf_city_filter)].copy() if pdf_city_filter else df_za_excel.copy()
                    wb = openpyxl.Workbook()

                    # ── Sheet 1: Pregled po gradovima ────────────────────────
                    ws_sum = wb.active
                    ws_sum.title = "Pregled po gradu"
                    header_fill  = PatternFill("solid", fgColor="009DE0")
                    total_fill   = PatternFill("solid", fgColor="1A1A2E")
                    green_fill   = PatternFill("solid", fgColor="D5F5E3")
                    red_fill     = PatternFill("solid", fgColor="FDECEA")
                    header_font  = Font(bold=True, color="FFFFFF")
                    total_font   = Font(bold=True, color="FFFFFF")
                    thin_border  = Border(
                        left=Side(style="thin", color="DDDDDD"),
                        right=Side(style="thin", color="DDDDDD"),
                        top=Side(style="thin", color="DDDDDD"),
                        bottom=Side(style="thin", color="DDDDDD"),
                    )

                    grad_sum_xl = fdf_xl.groupby("grad").agg(
                        Restorana=("naziv","count"),
                        Sa_promo=("akcije", lambda x: (x!="-").sum()),
                        Otvorenih=("status", lambda x: (x=="Otvoren").sum()),
                        Novih=("novo", lambda x: (x=="Da").sum()),
                    ).reset_index().sort_values("Sa_promo", ascending=False)
                    grad_sum_xl["Pct"] = (grad_sum_xl["Sa_promo"] / grad_sum_xl["Restorana"] * 100).round(1)

                    headers_s = ["Grad", "Ukupno rest.", "Sa promo", "% promo", "Otvorenih", "Novih"]
                    ws_sum.append(headers_s)
                    for cell in ws_sum[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center")
                        cell.border = thin_border

                    for _, r in grad_sum_xl.iterrows():
                        ws_sum.append([r["grad"], int(r["Restorana"]), int(r["Sa_promo"]),
                                       f"{r['Pct']}%", int(r["Otvorenih"]), int(r["Novih"])])
                        row_cells = ws_sum[ws_sum.max_row]
                        for c in row_cells:
                            c.border = thin_border
                            c.alignment = Alignment(horizontal="center")
                        # Boja prema promo procentu
                        pct_val = r["Pct"]
                        row_fill = green_fill if pct_val >= 30 else (red_fill if pct_val < 10 else None)
                        if row_fill:
                            for c in row_cells:
                                c.fill = row_fill

                    # Total red
                    total_row = ["UKUPNO", int(grad_sum_xl["Restorana"].sum()),
                                 int(grad_sum_xl["Sa_promo"].sum()),
                                 f"{round(grad_sum_xl['Sa_promo'].sum()/grad_sum_xl['Restorana'].sum()*100,1)}%",
                                 int(grad_sum_xl["Otvorenih"].sum()), int(grad_sum_xl["Novih"].sum())]
                    ws_sum.append(total_row)
                    for c in ws_sum[ws_sum.max_row]:
                        c.fill = total_fill
                        c.font = total_font
                        c.alignment = Alignment(horizontal="center")
                        c.border = thin_border

                    for col in ws_sum.columns:
                        ws_sum.column_dimensions[get_column_letter(col[0].column)].width = 16

                    # ── Sheet 2: Svi restorani ────────────────────────────────
                    ws_all = wb.create_sheet("Svi restorani")
                    disp_cols = ["grad","naziv","status","ocena","dostava","novo","akcije"]
                    disp_cols = [c for c in disp_cols if c in fdf_xl.columns]
                    col_names  = {"grad":"Grad","naziv":"Restoran","status":"Status","ocena":"Ocena",
                                  "dostava":"Dostava","novo":"Novo","akcije":"Akcije"}
                    ws_all.append([col_names.get(c,c) for c in disp_cols])
                    for cell in ws_all[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center")
                        cell.border = thin_border

                    for _, row_xl in fdf_xl[disp_cols].iterrows():
                        ws_all.append([str(row_xl[c]) if pd.notna(row_xl[c]) else "" for c in disp_cols])
                        r_cells = ws_all[ws_all.max_row]
                        for c in r_cells:
                            c.border = thin_border
                        has_promo = str(row_xl.get("akcije","-")) != "-"
                        if has_promo:
                            r_cells[disp_cols.index("akcije")].fill = green_fill if "akcije" in disp_cols else None

                    ws_all.column_dimensions["A"].width = 14
                    ws_all.column_dimensions["B"].width = 30
                    for i in range(3, len(disp_cols)+1):
                        ws_all.column_dimensions[get_column_letter(i)].width = 16

                    buf_xl = io.BytesIO()
                    wb.save(buf_xl)
                    st.download_button(
                        label="⬇️ Preuzmi Excel izveštaj",
                        data=buf_xl.getvalue(),
                        file_name=f"promo_izvestaj_{datetime.date.today().strftime('%Y%m%d')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="excel_download_btn",
                    )
                    st.success(f"✅ Excel generisan — {len(fdf_xl)} restorana u {fdf_xl['grad'].nunique()} gradova.")
                except ImportError:
                    st.error("❌ openpyxl nije instaliran. Pokreni: pip install openpyxl")
                except Exception as e:
                    st.error(f"❌ Greška pri generisanju Excel-a: {e}")

    if st.button("📄 Generiši PDF izveštaj", type="primary", key="pdf_gen_btn"):
        df_za_pdf = st.session_state.df_wolt
        if df_za_pdf.empty:
            st.warning("⚠️ Nema podataka skena. Najpre uradi scan.")
        else:
            try:
                from reportlab.lib.pagesizes import A4
                from reportlab.lib import colors
                from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                from reportlab.lib.units import cm
                from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                                Paragraph, Spacer, HRFlowable)
                import io

                buf = io.BytesIO()
                doc = SimpleDocTemplate(buf, pagesize=A4,
                                        leftMargin=1.5*cm, rightMargin=1.5*cm,
                                        topMargin=2*cm, bottomMargin=2*cm)
                styles = getSampleStyleSheet()
                WOLT_BLUE = colors.HexColor("#009de0")
                DARK      = colors.HexColor("#1a1a2e")
                GREEN     = colors.HexColor("#27ae60")
                ORANGE    = colors.HexColor("#e67e22")
                LIGHT_BG  = colors.HexColor("#f0f4ff")
                GREEN_BG  = colors.HexColor("#eafaf1")

                title_style = ParagraphStyle("title", parent=styles["Title"],
                    fontSize=20, textColor=DARK, spaceAfter=4)
                sub_style = ParagraphStyle("sub", parent=styles["Normal"],
                    fontSize=9, textColor=colors.grey, spaceAfter=10)
                h2_style = ParagraphStyle("h2", parent=styles["Heading2"],
                    fontSize=12, textColor=DARK, spaceBefore=16, spaceAfter=8)

                fdf = df_za_pdf[df_za_pdf["grad"].isin(pdf_city_filter)].copy() if pdf_city_filter else df_za_pdf.copy()

                story = []

                # ── Naslov ─────────────────────────────────────────────────
                story.append(Paragraph("Promo Monitor — Izveštaj", title_style))
                story.append(Paragraph(f"Generisano: {local_now()}", sub_style))
                story.append(HRFlowable(width="100%", thickness=2, color=WOLT_BLUE))
                story.append(Spacer(1, 0.4*cm))

                # ── Ukupni KPI ─────────────────────────────────────────────
                total_r  = len(fdf)
                total_p  = int((fdf["akcije"] != "-").sum())
                total_new = int((fdf["novo"] == "Da").sum()) if "novo" in fdf.columns else 0
                pct_p    = round(total_p / total_r * 100, 1) if total_r > 0 else 0
                n_cities = fdf["grad"].nunique()

                kpi_data = [
                    ["Ukupno restorana", "Sa promocijama", "% promo", "Novih", "Gradova"],
                    [str(total_r), str(total_p), f"{pct_p}%", str(total_new), str(n_cities)],
                ]
                kpi_table = Table(kpi_data, colWidths=[3.5*cm]*5)
                kpi_table.setStyle(TableStyle([
                    ("BACKGROUND",    (0,0), (-1,0),  DARK),
                    ("TEXTCOLOR",     (0,0), (-1,0),  colors.white),
                    ("FONTNAME",      (0,0), (-1,0),  "Helvetica-Bold"),
                    ("FONTSIZE",      (0,0), (-1,0),  8),
                    ("BACKGROUND",    (0,1), (-1,1),  LIGHT_BG),
                    ("FONTNAME",      (0,1), (-1,1),  "Helvetica-Bold"),
                    ("FONTSIZE",      (0,1), (-1,1),  16),
                    ("TEXTCOLOR",     (1,1), (1,1),   GREEN),
                    ("TEXTCOLOR",     (2,1), (2,1),   GREEN),
                    ("ALIGN",         (0,0), (-1,-1), "CENTER"),
                    ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
                    ("TOPPADDING",    (0,0), (-1,-1), 8),
                    ("BOTTOMPADDING", (0,0), (-1,-1), 8),
                    ("GRID",          (0,0), (-1,-1), 0.5, colors.HexColor("#dddddd")),
                ]))
                story.append(kpi_table)
                story.append(Spacer(1, 0.5*cm))

                # ── Statistika po gradu ─────────────────────────────────────
                story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#eeeeee")))
                story.append(Paragraph("Statistika po gradu", h2_style))

                grad_sum = fdf.groupby("grad").agg(
                    Restorana=("naziv", "count"),
                    Sa_promo=("akcije", lambda x: (x != "-").sum()),
                    Otvorenih=("status", lambda x: (x == "Otvoren").sum()),
                    Novih=("novo", lambda x: (x == "Da").sum()),
                ).reset_index().sort_values("Sa_promo", ascending=False)
                grad_sum["% promo"] = (grad_sum["Sa_promo"] / grad_sum["Restorana"] * 100).round(1).astype(str) + "%"

                city_data = [["Grad", "Ukupno rest.", "Sa promo", "% promo", "Otvorenih", "Novih"]]
                for _, r in grad_sum.iterrows():
                    city_data.append([
                        r["grad"],
                        str(int(r["Restorana"])),
                        str(int(r["Sa_promo"])),
                        r["% promo"],
                        str(int(r["Otvorenih"])),
                        str(int(r["Novih"])),
                    ])
                # Totali red
                city_data.append([
                    "UKUPNO",
                    str(int(grad_sum["Restorana"].sum())),
                    str(int(grad_sum["Sa_promo"].sum())),
                    f"{pct_p}%",
                    str(int(grad_sum["Otvorenih"].sum())),
                    str(int(grad_sum["Novih"].sum())),
                ])

                city_table = Table(city_data, colWidths=[4.5*cm, 2.8*cm, 2.5*cm, 2.2*cm, 2.5*cm, 2*cm])
                city_table.setStyle(TableStyle([
                    ("BACKGROUND",    (0,0),  (-1,0),  WOLT_BLUE),
                    ("TEXTCOLOR",     (0,0),  (-1,0),  colors.white),
                    ("FONTNAME",      (0,0),  (-1,0),  "Helvetica-Bold"),
                    ("FONTSIZE",      (0,0),  (-1,-1), 9),
                    ("ROWBACKGROUNDS",(0,1),  (-1,-2), [colors.white, LIGHT_BG]),
                    ("BACKGROUND",    (0,-1), (-1,-1), DARK),
                    ("TEXTCOLOR",     (0,-1), (-1,-1), colors.white),
                    ("FONTNAME",      (0,-1), (-1,-1), "Helvetica-Bold"),
                    ("GRID",          (0,0),  (-1,-1), 0.4, colors.HexColor("#dddddd")),
                    ("ALIGN",         (1,0),  (-1,-1), "CENTER"),
                    ("VALIGN",        (0,0),  (-1,-1), "MIDDLE"),
                    ("TOPPADDING",    (0,0),  (-1,-1), 6),
                    ("BOTTOMPADDING", (0,0),  (-1,-1), 6),
                ]))
                story.append(city_table)
                story.append(Spacer(1, 0.5*cm))

                # ── Statistika po AM-u ──────────────────────────────────────
                amm_df_curr = load_amm()
                if not amm_df_curr.empty:
                    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#eeeeee")))
                    story.append(Paragraph("Statistika po Account Manageru", h2_style))

                    fdf["naziv_norm"] = fdf["naziv"].apply(normalize)
                    merged_am = fdf.merge(
                        amm_df_curr[["restaurant_norm", "city", "am_name"]],
                        left_on=["naziv_norm", "grad"],
                        right_on=["restaurant_norm", "city"],
                        how="inner"
                    )

                    if not merged_am.empty:
                        am_sum = merged_am.groupby("am_name").agg(
                            Partnera=("naziv", "count"),
                            Sa_promo=("akcije", lambda x: (x != "-").sum()),
                            Gradova=("grad", "nunique"),
                        ).reset_index().sort_values("Sa_promo", ascending=False)
                        am_sum["% promo"] = (am_sum["Sa_promo"] / am_sum["Partnera"] * 100).round(1).astype(str) + "%"

                        am_data = [["Account Manager", "Partnera ukupno", "Sa promo", "% promo", "Gradova"]]
                        for _, r in am_sum.iterrows():
                            am_data.append([
                                r["am_name"],
                                str(int(r["Partnera"])),
                                str(int(r["Sa_promo"])),
                                r["% promo"],
                                str(int(r["Gradova"])),
                            ])

                        am_table = Table(am_data, colWidths=[5.5*cm, 3.5*cm, 2.5*cm, 2.5*cm, 2.5*cm])
                        am_table.setStyle(TableStyle([
                            ("BACKGROUND",    (0,0), (-1,0),  GREEN),
                            ("TEXTCOLOR",     (0,0), (-1,0),  colors.white),
                            ("FONTNAME",      (0,0), (-1,0),  "Helvetica-Bold"),
                            ("FONTSIZE",      (0,0), (-1,-1), 9),
                            ("ROWBACKGROUNDS",(0,1), (-1,-1), [colors.white, GREEN_BG]),
                            ("GRID",          (0,0), (-1,-1), 0.4, colors.HexColor("#dddddd")),
                            ("ALIGN",         (1,0), (-1,-1), "CENTER"),
                            ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
                            ("TOPPADDING",    (0,0), (-1,-1), 6),
                            ("BOTTOMPADDING", (0,0), (-1,-1), 6),
                        ]))
                        story.append(am_table)
                    else:
                        story.append(Paragraph("Nema podataka — partneri iz AM baze nisu pronađeni u skenu.", sub_style))

                # ── Footer ─────────────────────────────────────────────────
                story.append(Spacer(1, 0.8*cm))
                story.append(HRFlowable(width="100%", thickness=1, color=WOLT_BLUE))
                story.append(Spacer(1, 0.2*cm))
                story.append(Paragraph(
                    f"Promo Monitor — Automatski izveštaj — {local_now()}",
                    ParagraphStyle("footer", parent=styles["Normal"],
                                   fontSize=7, textColor=colors.grey, alignment=1)
                ))

                doc.build(story)
                pdf_bytes = buf.getvalue()

                st.download_button(
                    label="⬇️ Preuzmi PDF izveštaj",
                    data=pdf_bytes,
                    file_name=f"promo_izvestaj_{datetime.date.today().strftime('%Y%m%d')}.pdf",
                    mime="application/pdf",
                    key="pdf_download_btn",
                )
                st.success(f"✅ PDF generisan — {n_cities} gradova, {total_r} restorana, {total_p} sa promo.")
            except Exception as e:
                st.error(f"❌ Greška pri generisanju PDF-a: {e}")

# ══════════════════════════════════════════════════════════════════════════════
# TAB 5: AUTO-SCHEDULER
# ══════════════════════════════════════════════════════════════════════════════
with tab_sched:
    st.markdown("### ⏰ Automatic Scan Scheduler")
    cfg = load_scheduler_config()

    st.markdown("#### 🌍 Global Full Scan")
    st.caption("Runs a full scan across selected cities at a fixed daily time.")
    sc1, sc2, sc3 = st.columns(3)
    with sc1: sched_enabled = st.toggle("✅ Enable global scan", value=cfg.get("enabled", False))
    with sc2: sched_hour = st.number_input("Hour (0–23):", min_value=0, max_value=23, value=cfg.get("hour", 8))
    with sc3: sched_min = st.number_input("Minute (0–59):", min_value=0, max_value=59, value=cfg.get("minute", 0))
    sched_cities = st.multiselect("Cities:", options=CITIES, default=cfg.get("cities", CITIES))
    if st.button("💾 Save global schedule", type="primary"):
        new_cfg = dict(cfg)
        new_cfg.update({"enabled": sched_enabled, "hour": int(sched_hour), "minute": int(sched_min), "cities": sched_cities})
        save_scheduler_config(new_cfg)
        st.success(f"✅ Saved! Global scan {'ENABLED' if sched_enabled else 'DISABLED'} at **{int(sched_hour):02d}:{int(sched_min):02d}**.")

    st.markdown("---")

    st.markdown("#### 🏙️ Per-City Schedules")
    st.caption("Set individual scan times and days for each city. These run City Rescan (merges with existing data).")

    DAYS = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
    DAY_FULL = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]

    city_schedules = cfg.get("city_schedules", {})

    for city in CITIES:
        city_cfg = city_schedules.get(city, {"enabled": False, "hour": 8, "minute": 0, "days": []})
        with st.expander(f"🏙️ {city}", expanded=city_cfg.get("enabled", False)):
            cc1, cc2, cc3 = st.columns(3)
            with cc1:
                city_enabled = st.toggle(f"Enable for {city}", value=city_cfg.get("enabled", False), key=f"city_en_{city}")
            with cc2:
                city_hour = st.number_input("Hour (0–23):", min_value=0, max_value=23,
                                             value=city_cfg.get("hour", 8), key=f"city_h_{city}")
            with cc3:
                city_min = st.number_input("Minute (0–59):", min_value=0, max_value=59,
                                            value=city_cfg.get("minute", 0), key=f"city_m_{city}")

            st.markdown("**Days of week:**")
            day_cols = st.columns(7)
            selected_days = []
            for di, (short, full) in enumerate(zip(DAYS, DAY_FULL)):
                with day_cols[di]:
                    checked = st.checkbox(short, value=(short in city_cfg.get("days", [])),
                                          key=f"city_day_{city}_{short}", help=full)
                    if checked:
                        selected_days.append(short)

            if st.button(f"💾 Save {city}", key=f"city_save_{city}"):
                new_cfg = dict(cfg)
                if "city_schedules" not in new_cfg:
                    new_cfg["city_schedules"] = {}
                new_cfg["city_schedules"][city] = {
                    "enabled": city_enabled,
                    "hour": int(city_hour),
                    "minute": int(city_min),
                    "days": selected_days,
                }
                save_scheduler_config(new_cfg)
                cfg = load_scheduler_config()
                city_schedules = cfg.get("city_schedules", {})
                days_str = ", ".join(selected_days) if selected_days else "no days selected"
                st.success(f"✅ {city}: {'ENABLED' if city_enabled else 'DISABLED'} at **{int(city_hour):02d}:{int(city_min):02d}** on {days_str}")

    st.markdown("---")

    st.markdown("#### 🧪 Test – run manually now")
    sched_running = st.session_state.get("sched_running", False)
    sched_done    = st.session_state.get("sched_done", False)
    if st.button("▶️ Run test scan + send now", disabled=sched_running):
        st.session_state["sched_running"] = True
        st.session_state["sched_done"]    = False
        def _run_sched_bg():
            run_scheduled_scan_and_send()
            st.session_state["sched_running"] = False
            st.session_state["sched_done"]    = True
        threading.Thread(target=_run_sched_bg, daemon=True).start()
        st.rerun()
    if sched_running:
        st.info("🔄 Running...")
        time.sleep(3)
        st.rerun()
    if sched_done:
        st.session_state["sched_done"] = False
        st.success("✅ Test finished.")

    st.markdown("---")

    cfg_cur = load_scheduler_config()
    if cfg_cur.get("enabled"):
        now = datetime.datetime.now()
        target = now.replace(hour=cfg_cur["hour"], minute=cfg_cur["minute"], second=0, microsecond=0)
        if now >= target:
            target += datetime.timedelta(days=1)
        diff = target - now
        h, rem = divmod(int(diff.total_seconds()), 3600)
        m_rem = rem // 60
        st.success(f"🕐 Next global scan in: **{h}h {m_rem}min** (at {cfg_cur['hour']:02d}:{cfg_cur['minute']:02d})")
    else:
        st.warning("Global automatic scan is disabled.")

    city_scheds_cur = cfg_cur.get("city_schedules", {})
    active_city_scheds = {c: s for c, s in city_scheds_cur.items() if s.get("enabled") and s.get("days")}
    if active_city_scheds:
        st.markdown("**Active per-city schedules:**")
        for city, cs in active_city_scheds.items():
            days_str = ", ".join(cs["days"])
            st.info(f"🏙️ **{city}** — {cs['hour']:02d}:{cs['minute']:02d} on {days_str}")

# ══════════════════════════════════════════════════════════════════════════════
# ══════════════════════════════════════════════════════════════════════════════
# TAB 6: WATCHLIST
# ══════════════════════════════════════════════════════════════════════════════
with tab_watchlist:
    st.markdown("### ⭐ Watchlist — Praćeni Restorani")
    st.caption("Dodaj restorane koje želiš prioritetno pratiti. Prikazuju se odvojeno sa trenutnim akcijama, bez cooldowna i bez potrebe za AM bazom.")

    df_wolt_wl = st.session_state.df_wolt
    wl_df = load_watchlist()

    # ── Dodavanje na watchlistu ───────────────────────────────────────────────
    st.markdown("#### ➕ Dodaj na Watchlist")
    wl_c1, wl_c2, wl_c3 = st.columns([2, 1, 1])
    with wl_c1:
        if not df_wolt_wl.empty and "naziv" in df_wolt_wl.columns:
            rest_options_wl = sorted(df_wolt_wl["naziv"].dropna().unique().tolist())
            wl_sel = st.selectbox("Izaberi iz skena:", ["-- Izaberi --"] + rest_options_wl, key="wl_sel_rest")
        else:
            wl_sel = "-- Izaberi --"
            st.info("Pokreni sken da bi dodavao restorane.")
    with wl_c2:
        wl_city_sel = st.selectbox("Grad:", ["-- Svi --"] + CITIES, key="wl_city_sel")
    with wl_c3:
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("⭐ Dodaj na Watchlist", key="wl_add_btn"):
            if wl_sel and wl_sel != "-- Izaberi --":
                city_val = "" if wl_city_sel == "-- Svi --" else wl_city_sel
                # Pronađi slug
                mask_wl = df_wolt_wl["naziv"] == wl_sel
                if city_val:
                    mask_wl = mask_wl & (df_wolt_wl["grad"] == city_val)
                row_wl = df_wolt_wl[mask_wl]
                if not row_wl.empty and "slug" in row_wl.columns:
                    slug_wl = row_wl.iloc[0]["slug"]
                    grad_wl = row_wl.iloc[0]["grad"]
                    add_to_watchlist(slug_wl, grad_wl, wl_sel)
                    st.success(f"✅ **{wl_sel}** ({grad_wl}) dodat na watchlist!")
                    st.rerun()
                else:
                    st.error("Restoran nije pronađen u skenu.")
            else:
                st.warning("Izaberi restoran.")

    # Manuelno dodavanje po slugu
    with st.expander("➕ Dodaj manuelno po slug-u"):
        mc1, mc2, mc3, mc4 = st.columns([2, 1, 1, 1])
        with mc1: man_slug = st.text_input("Slug:", placeholder="mcdonalds-beograd", key="wl_man_slug")
        with mc2: man_city = st.selectbox("Grad:", CITIES, key="wl_man_city")
        with mc3: man_naziv = st.text_input("Naziv:", placeholder="McDonald's", key="wl_man_naziv")
        with mc4:
            st.markdown("<br>", unsafe_allow_html=True)
            if st.button("Dodaj", key="wl_man_add"):
                if man_slug and man_naziv:
                    add_to_watchlist(man_slug, man_city, man_naziv)
                    st.success(f"✅ Dodat: {man_naziv}")
                    st.rerun()

    st.markdown("---")

    # ── Prikaz watchliste sa aktuelnim akcijama ───────────────────────────────
    st.markdown("#### 📋 Trenutna Watchlista")

    wl_df = load_watchlist()
    if wl_df.empty:
        st.info("Watchlista je prazna. Dodaj restorane gore.")
    else:
        # Spoji sa podacima iz skena
        if not df_wolt_wl.empty and "slug" in df_wolt_wl.columns:
            wl_enriched = wl_df.merge(
                df_wolt_wl[["slug","grad","akcije","status","link"]].rename(columns={"grad":"grad_scan"}),
                on="slug", how="left"
            )
            wl_enriched["akcije"]  = wl_enriched["akcije"].fillna("⏳ Sken nije pokrenut")
            wl_enriched["status"]  = wl_enriched["status"].fillna("—")
            wl_enriched["link"]    = wl_enriched["link"].fillna("")
        else:
            wl_enriched = wl_df.copy()
            wl_enriched["akcije"] = "⏳ Sken nije pokrenut"
            wl_enriched["status"] = "—"
            wl_enriched["link"]   = ""

        # Prikaži kartice za restorane sa promocijama
        has_promo_wl = wl_enriched[wl_enriched["akcije"].apply(
            lambda x: x not in ["-","⏳ Sken nije pokrenut",""] and pd.notna(x)
        )]
        no_promo_wl  = wl_enriched[~wl_enriched.index.isin(has_promo_wl.index)]

        if not has_promo_wl.empty:
            st.markdown(f"**🟢 Sa aktivnim promocijama ({len(has_promo_wl)}):**")
            for _, r_wl in has_promo_wl.iterrows():
                akcije_html = r_wl["akcije"].replace("\n","<br>") if r_wl["akcije"] != "-" else "—"
                link_html = f"<a href='{r_wl['link']}' target='_blank' style='color:#009de0;font-size:12px'>Otvori ↗</a>" if r_wl["link"] else ""
                st.markdown(f"""
                <div style='background:#fff;border-radius:10px;padding:14px 20px;
                            box-shadow:0 2px 8px rgba(0,0,0,0.07);border-left:4px solid #27ae60;
                            margin-bottom:10px'>
                  <div style='display:flex;justify-content:space-between;align-items:center'>
                    <div>
                      <span style='font-weight:800;font-size:1rem'>{r_wl['naziv']}</span>
                      <span style='color:#888;font-size:0.85rem;margin-left:10px'>{r_wl.get('city','') or r_wl.get('grad_scan','')}</span>
                      &nbsp;{link_html}
                    </div>
                    <span style='font-size:0.8rem;color:#27ae60;font-weight:700'>{r_wl.get('status','')}</span>
                  </div>
                  <div style='margin-top:8px;font-size:0.88rem;color:#333'>{akcije_html}</div>
                </div>""", unsafe_allow_html=True)

        if not no_promo_wl.empty:
            with st.expander(f"⚪ Bez aktivnih promocija ({len(no_promo_wl)})", expanded=False):
                for _, r_wl in no_promo_wl.iterrows():
                    st.markdown(f"**{r_wl['naziv']}** — {r_wl.get('city','') or r_wl.get('grad_scan','')} — _{r_wl['akcije']}_")

        st.markdown("---")
        st.markdown("**Upravljanje watchlistom:**")
        wl_display = wl_df[["naziv","city","slug","added_at"]].copy()
        wl_display.columns = ["Naziv","Grad","Slug","Dodat"]
        st.dataframe(wl_display, use_container_width=True, hide_index=True, height=200)

        rm_c1, rm_c2 = st.columns([2, 1])
        with rm_c1:
            rm_options = [f"{r['naziv']} ({r['city']})" for _, r in wl_df.iterrows()]
            rm_sel = st.selectbox("Ukloni sa watchliste:", ["-- Izaberi --"] + rm_options, key="wl_rm_sel")
        with rm_c2:
            st.markdown("<br>", unsafe_allow_html=True)
            if st.button("🗑️ Ukloni", key="wl_rm_btn") and rm_sel != "-- Izaberi --":
                idx = rm_options.index(rm_sel)
                row_rm = wl_df.iloc[idx]
                remove_from_watchlist(row_rm["slug"], row_rm["city"])
                st.success(f"Uklonjen: {row_rm['naziv']}")
                st.rerun()

        st.markdown("---")

        # ── Instant alert dugme za watchlistu ─────────────────────────────────
        st.markdown("#### 📧 Instant Alert za Watchlist")
        st.caption("Šalje email/Slack za sve restorane sa watchliste koji trenutno imaju promocije — bez cooldowna.")
        wl_alert_email = st.text_input("Pošalji na email:", placeholder="tvoj@email.com", key="wl_alert_email")
        if st.button("📧 Pošalji Watchlist Alert", type="primary", key="wl_alert_btn"):
            if not wl_alert_email:
                st.error("Unesi email adresu.")
            elif has_promo_wl.empty:
                st.info("Nema restorana sa aktivnim promocijama na watchlisti.")
            else:
                alerts_wl = has_promo_wl[["naziv","city","akcije","link"]].rename(
                    columns={"naziv":"naziv","city":"grad","akcije":"akcije","link":"link"}
                ).to_dict("records")
                ok_wl = send_alert_email(wl_alert_email, "Watchlist", alerts_wl, tip="Watchlist")
                if ok_wl:
                    st.success(f"✅ Watchlist alert poslat na {wl_alert_email} — {len(alerts_wl)} restorana!")
                else:
                    st.error("❌ Greška pri slanju. Proveri email konfiguraciju.")

# ══════════════════════════════════════════════════════════════════════════════
# TAB 7: DEBUG API (prethodno TAB 6)
# ══════════════════════════════════════════════════════════════════════════════
with tab_debug:
    st.markdown("### ⚙️ Settings & Debug")

    _debug_pass = st.text_input("🔑 Password to access:", type="password", key="debug_pass_input")
    if _debug_pass != SETTINGS_PASSWORD:
        st.warning("Enter the password to access Settings.")
        st.stop()

    st.markdown("#### 📊 Lokalna SQLite Baza Status")
    try:
        with sqlite3.connect(DB_FILE, check_same_thread=False) as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
            tables = [t[0] for t in cursor.fetchall()]
        st.success(f"✅ Lokalna SQLite baza OK — Fajl: **{DB_FILE}**")
        st.info(f"Tabele u bazi: {', '.join(tables)}")
    except Exception as e:
        st.error(f"❌ Greška sa lokalnom bazom: {e}")

    st.markdown("---")
    st.markdown("#### 💬 Slack Webhook Notifikacije")
    st.caption("Kada se šalje AM alert mejl, ista informacija ide i na Slack kanal. Ostavi prazno da deaktiviraj.")
    current_slack = load_slack_webhook()
    new_slack = st.text_input(
        "Slack Webhook URL:",
        value=current_slack,
        placeholder="https://hooks.slack.com/services/...",
        key="slack_webhook_input",
        type="password" if current_slack else "default",
    )
    slack_col1, slack_col2 = st.columns([1, 3])
    with slack_col1:
        if st.button("💾 Sačuvaj Slack URL", key="slack_save"):
            save_slack_webhook(new_slack)
            st.success("✅ Slack webhook sačuvan!")
    with slack_col2:
        if st.button("🧪 Test Slack poruka", key="slack_test") and new_slack:
            fake_alerts = [{"naziv": "Test Restoran", "grad": "Beograd", "akcije": "• 20% popust na sve"}]
            ok_sl = send_slack_notification(new_slack, fake_alerts, "Test AM")
            if ok_sl:
                st.success("✅ Test Slack poruka uspešno poslata!")
            else:
                st.error("❌ Greška. Proveri webhook URL.")
    if current_slack:
        st.info(f"✅ Slack webhook je aktivan.")
    else:
        st.warning("Slack webhook nije konfigurisan — alertovi idu samo na email.")

    st.markdown("---")
    st.markdown("#### 📧 Test Email Slanje")
    st.caption("Pošalji test mail sa lažnim podacima da proveriš da li SMTP radi i kako izgleda forma.")

    # ── Auto-popunjavanje imena iz email adrese ───────────────────────────────
    def _name_from_email(email: str) -> str:
        """
        veber.zlatan@glovoapp.com  →  Veber Zlatan
        zlatan.veber@firma.com     →  Zlatan Veber
        zlatanveber@firma.com      →  Zlatanveber  (nema tačke → ne može da razdvoji)
        """
        if not email or "@" not in email:
            return ""
        local = email.split("@")[0]          # deo pre @
        parts = local.replace("_", ".").split(".")
        return " ".join(p.capitalize() for p in parts if p)

    test_email_addr = st.text_input(
        "📬 Pošalji test na email:",
        placeholder="ime.prezime@firma.com",
        key="test_email_input"
    )

    # Ako je email unet, pokušaj auto-fill imena
    auto_name = _name_from_email(test_email_addr) if test_email_addr else ""
    test_am_name = st.text_input(
        "👤 Ime AM-a (za AM test mail):",
        value=auto_name,
        placeholder="Automatski iz emaila ili unesi ručno",
        key="test_am_name"
    )
    if test_email_addr and auto_name and auto_name != "Marko Marković":
        st.caption(f"💡 Ime automatski preuzeto iz email adrese: **{auto_name}**")

    tbtn1, tbtn2 = st.columns(2)
    with tbtn1:
        if st.button("📤 Pošalji TEST mail za AM-a", key="test_am_btn", type="primary"):
            if not test_email_addr:
                st.error("Unesi email adresu.")
            else:
                # Novi restorani sa promo (prikazuju se kao "alert" — gornji deo maila)
                fake_alerts = [
                    {
                        "naziv": "Burger House Beograd",
                        "grad": "Beograd",
                        "akcije": "• 20% popust na sve burgere\n• Besplatna dostava za narudžbine preko 1500 din",
                        "link": "https://wolt.com/sr/srb/belgrade/restaurant/burger-house",
                        "norm": "burger house beograd",
                    },
                    {
                        "naziv": "Pizza Roma",
                        "grad": "Niš",
                        "akcije": "• [Wolt+] 2+1 na sve pizze\n• Gratis napitak uz svaku narudžbinu",
                        "link": "https://wolt.com/sr/srb/nis/restaurant/pizza-roma",
                        "norm": "pizza roma",
                    },
                    {
                        "naziv": "Sushi Bar Zen",
                        "grad": "Novi Sad",
                        "akcije": "• 15% popust vikendima",
                        "link": "https://wolt.com/sr/srb/novi-sad/restaurant/sushi-bar-zen",
                        "norm": "sushi bar zen",
                    },
                ]
                # Svi partneri AM-a (prikazuju se u donjem delu maila kao kompletan spisak)
                fake_all_partners = [
                    {
                        "naziv": "Burger House Beograd",
                        "grad": "Beograd",
                        "akcije": "• 20% popust na sve burgere\n• Besplatna dostava za narudžbine preko 1500 din",
                        "link": "https://wolt.com/sr/srb/belgrade/restaurant/burger-house",
                    },
                    {
                        "naziv": "Pizza Roma",
                        "grad": "Niš",
                        "akcije": "• [Wolt+] 2+1 na sve pizze\n• Gratis napitak uz svaku narudžbinu",
                        "link": "https://wolt.com/sr/srb/nis/restaurant/pizza-roma",
                    },
                    {
                        "naziv": "Sushi Bar Zen",
                        "grad": "Novi Sad",
                        "akcije": "• 15% popust vikendima",
                        "link": "https://wolt.com/sr/srb/novi-sad/restaurant/sushi-bar-zen",
                    },
                    {
                        "naziv": "Taco Fiesta",
                        "grad": "Beograd",
                        "akcije": "-",
                        "link": "https://wolt.com/sr/srb/belgrade/restaurant/taco-fiesta",
                    },
                    {
                        "naziv": "Green Bowl",
                        "grad": "Beograd",
                        "akcije": "-",
                        "link": "https://wolt.com/sr/srb/belgrade/restaurant/green-bowl",
                    },
                    {
                        "naziv": "Smokehouse BBQ",
                        "grad": "Novi Sad",
                        "akcije": "-",
                        "link": "https://wolt.com/sr/srb/novi-sad/restaurant/smokehouse-bbq",
                    },
                ]
                ok = send_alert_email(
                    test_email_addr, test_am_name,
                    fake_alerts,
                    all_partners=fake_all_partners,
                    tip="AM-Test"
                )
                if ok:
                    st.success(f"✅ Test AM mail uspešno poslat na **{test_email_addr}**!")
                else:
                    st.error("❌ Slanje nije uspelo. Proveri EMAIL_SENDER i EMAIL_PASSWORD u .env fajlu.")

    with tbtn2:
        if st.button("📤 Pošalji TEST mail za Prodavca", key="test_sales_btn", type="primary"):
            if not test_email_addr:
                st.error("Unesi email adresu.")
            else:
                fake_novi = [
                    {"naziv": "Taco Loco", "slug": "taco-loco"},
                    {"naziv": "Green Bowl", "slug": "green-bowl"},
                    {"naziv": "Smokehouse BBQ", "slug": "smokehouse-bbq"},
                ]
                ok = send_sales_bulk_notification(test_email_addr, "Beograd", fake_novi)
                if ok:
                    st.success(f"✅ Test Sales mail uspešno poslat na **{test_email_addr}**!")
                else:
                    st.error("❌ Slanje nije uspelo. Proveri EMAIL_SENDER i EMAIL_PASSWORD u .env fajlu.")

    st.markdown("---")
    if "custom_coords" not in st.session_state:
        st.session_state["custom_coords"] = {k: list(v) for k, v in CITY_MULTI_COORDS.items()}

    for city_key in CITY_KEYS:
        city_disp = CITY_DISPLAY.get(city_key, city_key)
        coords_list = st.session_state["custom_coords"].get(city_key, [])
        with st.expander(f"📍 {city_disp} — {len(coords_list)} locations", expanded=False):
            coords_text = "\n".join(f"{lat}, {lon}" for lat, lon in coords_list)
            new_text = st.text_area(f"Coordinates for {city_disp}:", value=coords_text,
                                    height=max(120, len(coords_list) * 35 + 40),
                                    key=f"coords_input_{city_key}", label_visibility="collapsed")
            col_save_c, col_reset_c = st.columns(2)
            with col_save_c:
                if st.button("💾 Save coordinates", key=f"save_coords_{city_key}"):
                    parsed_coords = []
                    errors = []
                    for i, line in enumerate(new_text.strip().split("\n")):
                        line = line.strip()
                        if not line:
                            continue
                        try:
                            parts = [p.strip().replace(",", ".") for p in line.replace(";", ",").split(",")]
                            if len(parts) >= 2:
                                lat_v, lon_v = float(parts[0]), float(parts[1])
                                if -90 <= lat_v <= 90 and -180 <= lon_v <= 180:
                                    parsed_coords.append((lat_v, lon_v))
                                else:
                                    errors.append(f"Row {i+1}: out of range")
                            else:
                                errors.append(f"Row {i+1}: format must be `lat, lon`")
                        except ValueError:
                            errors.append(f"Row {i+1}: not a number")
                    if errors:
                        for err in errors: st.error(err)
                    elif not parsed_coords:
                        st.error("No valid coordinates.")
                    else:
                        st.session_state["custom_coords"][city_key] = parsed_coords
                        st.success(f"✅ {len(parsed_coords)} locations saved.")
            with col_reset_c:
                if st.button("↩️ Reset to default", key=f"reset_coords_{city_key}"):
                    st.session_state["custom_coords"][city_key] = list(CITY_MULTI_COORDS[city_key])
                    st.success("↩️ Reset.")
                    st.rerun()

    st.markdown("---")
    st.markdown("### 🔬 Raw API response")
    dc1, dc2 = st.columns([2, 1])
    with dc1: debug_slug = st.text_input("Restaurant slug:", placeholder="e.g. mcdonalds-nis", key="debug_slug")
    with dc2: debug_city_display = st.selectbox("City:", CITIES, key="debug_city")

    if st.button("🔍 Fetch JSON", key="debug_fetch") and debug_slug:
        debug_city_key = display_to_key(debug_city_display)
        lat, lon  = CITY_COORDS.get(debug_city_key, (44.8178, 20.4569))
        city_slug = CITY_SLUG_MAP.get(debug_city_key, "belgrade")
        dyn_url = (f"https://consumer-api.wolt.com/order-xp/web/v1/venue/slug/{debug_slug}/dynamic/"
                   f"?lat={lat}&lon={lon}&selected_delivery_method=homedelivery")
        dyn_data, dyn_status = wolt_get(dyn_url)
        if dyn_data:
            with st.expander("Full JSON", expanded=True):
                st.json(dyn_data)
            parsed = _parse_dynamic_with_item_discount(dyn_data)
            st.markdown("**Parsed promotions:**")
            for p in parsed: st.write(p)
            if not parsed: st.warning("No parsed promotions.")
        else:
            st.warning(f"No data returned. HTTP: {dyn_status}")

    st.markdown("---")
    st.markdown("### 📋 Fetch Debug Log")
    col_log1, col_log2 = st.columns(2)
    with col_log1:
        if st.button("🔄 Refresh log"): st.rerun()
    with col_log2:
        if st.button("🗑️ Delete log"):
            Path("_fetch_debug.log").unlink(missing_ok=True)
            st.success("Log deleted.")
    try:
        log_content = Path("_fetch_debug.log").read_text(encoding="utf-8")
        if log_content.strip():
            lines = log_content.strip().split("\n")
            st.code("\n".join(lines[-200:]), language=None)
        else:
            st.info("Log is empty.")
    except FileNotFoundError:
        st.info("Log does not exist.")

# ══════════════════════════════════════════════════════════════════════════════
# TAB 8: RESET & BACKUP
# ══════════════════════════════════════════════════════════════════════════════
with tab_reset:
    st.markdown("### 🗑️ Reset System")
    RESET_PASSWORD = RESET_PASSWORD_ENV

    st.markdown("#### 💾 Backup")
    if st.button("📦 Create backup (CSV download)", key="backup_btn"):
        df_wolt_bk = st.session_state.df_wolt
        if not df_wolt_bk.empty:
            st.download_button("⬇️ Download scan CSV", df_wolt_bk.to_csv(index=False).encode("utf-8"),
                               file_name=f"scan_backup_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.csv",
                               mime="text/csv", key="backup_scan_dl")
        amm_bk = load_amm()
        if not amm_bk.empty:
            st.download_button("⬇️ Download AM CSV", amm_bk.to_csv(index=False).encode("utf-8"),
                               file_name="amm_backup.csv", mime="text/csv", key="backup_amm_dl")

    st.markdown("---")
    st.markdown("#### ⚠️ Reset operations")
    reset_pass = st.text_input("🔑 Password:", type="password", key="reset_pass_input")
    pass_ok = reset_pass == RESET_PASSWORD

    r1, r2, r3, r4 = st.columns(4)
    with r1:
        st.markdown("**Reset logs**")
        if st.button("🗑️ Delete logs", key="reset_logs", disabled=not pass_ok):
            Path("_fetch_debug.log").unlink(missing_ok=True)
            st.success("✅ Logs deleted.")
    with r2:
        st.markdown("**Reset AM database**")
        if st.button("🗑️ Delete AM database", key="reset_amm", disabled=not pass_ok):
            empty_amm = pd.DataFrame(columns=AMM_COLS)
            save_amm(empty_amm)
            st.success("✅ AM database deleted.")
    with r3:
        st.markdown("**Reset scan results**")
        if st.button("🗑️ Delete scan", key="reset_scan", disabled=not pass_ok):
            st.session_state.df_wolt = pd.DataFrame()
            st.session_state.last_scan = None
            st.session_state.scan_duration_last = None
            SCAN_FILE.unlink(missing_ok=True)
            Path("_scan_result.json").unlink(missing_ok=True)
            save_scan_gsheet(pd.DataFrame())
            st.success("✅ Scan deleted.")
            st.rerun()
    with r4:
        st.markdown("**Reset EVERYTHING**")
        if st.button("💥 RESET ALL", key="reset_all", type="primary", disabled=not pass_ok):
            st.session_state.df_wolt = pd.DataFrame()
            st.session_state.last_scan = None
            for f in ["_scan_result.json", "_scan_done.txt", "_scan_status.txt", "_scan_city_progress.json", "_fetch_debug.log", "promo_monitor.db"]:
                Path(f).unlink(missing_ok=True)
            SCAN_FILE.unlink(missing_ok=True)
            save_amm(pd.DataFrame(columns=AMM_COLS))
            save_alert_log_gsheet(pd.DataFrame(columns=ALERT_COLS))
            # Re-init baze da se obnovi promo_state i ostale tabele
            init_db()
            st.success("💥 Everything deleted!")
            st.rerun()

    if reset_pass and not pass_ok:
        st.error("❌ Wrong password.")
